import streamlit as st
from datetime import datetime, date
import base64
import sqlite3
import pandas as pd
import hashlib
import math
from pathlib import Path
from html import escape
from io import BytesIO
from PIL import Image, ImageDraw, ImageOps, ImageFont
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
import tempfile
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from st_aggrid import AgGrid, GridOptionsBuilder

BASE_DIR = Path(__file__).resolve().parent
DB_PATH = BASE_DIR / "cases.db"

try:
    import streamlit_drawable_canvas as drawable_canvas
    from streamlit_drawable_canvas import st_canvas
except ImportError:
    drawable_canvas = None
    st_canvas = None


def patch_streamlit_drawable_canvas_image_to_url():
    try:
        import streamlit.elements.image as st_image
        from streamlit.elements.lib.image_utils import image_to_url
        from streamlit.elements.lib.layout_utils import LayoutConfig
    except Exception:
        return

    def image_to_url_compat(
        image,
        width,
        clamp,
        channels,
        output_format,
        image_id,
    ):
        try:
            image_url = image_to_url(
                image,
                layout_config=LayoutConfig(width=width),
                clamp=clamp,
                channels=channels,
                output_format=output_format,
                image_id=image_id,
            )
            if (
                image_url
                and not image_url.startswith(("/", "http://", "https://", "data:"))
            ):
                image_url = f"/{image_url}"
            return image_url
        except TypeError:
            return image_to_url(
                image,
                width,
                clamp,
                channels,
                output_format,
                image_id,
            )

    st_image.image_to_url = image_to_url_compat


patch_streamlit_drawable_canvas_image_to_url()


def get_canvas_background_image_url(image, width, image_id):
    from streamlit.elements.lib.image_utils import image_to_url
    from streamlit.elements.lib.layout_utils import LayoutConfig

    image_buffer = BytesIO()
    image = ImageOps.exif_transpose(image).convert("RGB")
    image.save(image_buffer, format="PNG")
    image_buffer.seek(0)

    image_url = image_to_url(
        image_buffer,
        layout_config=LayoutConfig(width=width),
        clamp=True,
        channels="RGB",
        output_format="PNG",
        image_id=image_id,
    )

    if not image_url:
        return None

    base_url_path = st._config.get_option("server.baseUrlPath") or ""
    if base_url_path and not base_url_path.startswith("/"):
        base_url_path = f"/{base_url_path}"
    base_url_path = base_url_path.rstrip("/")

    if image_url.startswith(("http://", "https://", "data:")):
        return image_url

    if not image_url.startswith("/"):
        image_url = f"/{image_url}"

    if base_url_path and not image_url.startswith(f"{base_url_path}/"):
        image_url = f"{base_url_path}{image_url}"

    return image_url


def patch_streamlit_drawable_canvas_background_url():
    global st_canvas

    if drawable_canvas is None:
        return

    def st_canvas_with_fixed_background_url(
        fill_color="#eee",
        stroke_width=20,
        stroke_color="black",
        background_color="",
        background_image=None,
        update_streamlit=True,
        height=400,
        width=600,
        drawing_mode="freedraw",
        initial_drawing=None,
        display_toolbar=True,
        point_display_radius=3,
        key=None,
    ):
        background_image_url = None
        if background_image:
            background_image = drawable_canvas._resize_img(
                background_image,
                height,
                width,
            )
            background_image_url = get_canvas_background_image_url(
                background_image,
                width,
                (
                    "drawable-canvas-bg-"
                    f"{hashlib.md5(background_image.tobytes()).hexdigest()}-{key}"
                ),
            )
            background_color = ""

        initial_drawing = (
            {"version": "4.4.0"} if initial_drawing is None else initial_drawing
        )
        initial_drawing["background"] = background_color

        component_value = drawable_canvas._component_func(
            fillColor=fill_color,
            strokeWidth=stroke_width,
            strokeColor=stroke_color,
            backgroundColor=background_color,
            backgroundImageURL=background_image_url,
            realtimeUpdateStreamlit=update_streamlit and (drawing_mode != "polygon"),
            canvasHeight=height,
            canvasWidth=width,
            drawingMode=drawing_mode,
            initialDrawing=initial_drawing,
            displayToolbar=display_toolbar,
            displayRadius=point_display_radius,
            key=key,
            default=None,
        )

        if component_value is None:
            return drawable_canvas.CanvasResult()

        return drawable_canvas.CanvasResult(
            drawable_canvas.np.asarray(
                drawable_canvas._data_url_to_image(component_value["data"])
            ),
            component_value["raw"],
        )

    drawable_canvas.st_canvas = st_canvas_with_fixed_background_url
    st_canvas = st_canvas_with_fixed_background_url


patch_streamlit_drawable_canvas_background_url()

DETAIL_PHOTO_PREVIEW_SIZE = (980, 310)
MARK_CANVAS_MAX_WIDTH = 520
MARK_CANVAS_RENDER_VERSION = 3
CASE_CATEGORIES = ["파손", "오발송", "쇼트", "불량", "누락", "기타", "수평", "용접"]
CASE_EXCEL_FORM_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
CASE_EXCEL_FIELD_CELLS = {
    "case_date": "B3",
    "category": "B4",
    "product": "B5",
    "barcode": "B6",
    "cause": "B11",
    "action": "B18",
    "repair_method": "B21",
    "prevention": "B28",
}
CASE_EXCEL_IMAGE_RANGES = {
    "product_image": (7, 10),
    "case_image": (14, 17),
    "repair_image": (24, 27),
}
MARK_COLORS = {
    "빨강": "#ff2d2d",
    "노랑": "#ffd400",
    "초록": "#31d158",
    "파랑": "#0a84ff",
}


def fit_detail_photo(img):
    img = ImageOps.exif_transpose(img)
    return ImageOps.contain(
        img,
        DETAIL_PHOTO_PREVIEW_SIZE,
        method=Image.LANCZOS,
    )


def get_first_upload_bytes(uploaded_file):
    if not uploaded_file:
        return None

    if isinstance(uploaded_file, list):
        if len(uploaded_file) == 0:
            return None
        return uploaded_file[0].getvalue()

    return uploaded_file.getvalue()


def normalize_excel_text(value):
    if value is None:
        return ""
    return str(value).strip()


def parse_case_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value:
        text = str(value).strip()
        for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d", "%Y%m%d"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                pass
    return datetime.now().date()


def create_case_excel_form(default_case_date=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "사례등록"

    default_case_date = default_case_date or datetime.now().date()
    thin_gray = Side(style="thin", color="D9E2EC")
    border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
    label_fill = PatternFill("solid", fgColor="EAF4F2")
    photo_fill = PatternFill("solid", fgColor="F8FAFC")
    title_font = Font(size=16, bold=True, color="0F172A")
    label_font = Font(size=11, bold=True, color="0F172A")
    input_font = Font(size=11, color="0F172A")

    ws.merge_cells("A1:D1")
    ws["A1"] = "사례 등록 엑셀 폼"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 28

    fields = [
        (3, "발생일자", default_case_date.strftime("%Y-%m-%d")),
        (4, "유형", "파손"),
        (5, "상품명", ""),
        (6, "바코드", ""),
        (11, "원인", ""),
        (18, "조치방법", ""),
        (21, "수리방법", ""),
        (28, "방지대책", ""),
    ]

    for row, label, value in fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = label_font
        ws[f"A{row}"].fill = label_fill
        ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"A{row}"].border = border
        ws[f"B{row}"] = value
        ws[f"B{row}"].font = input_font
        ws[f"B{row}"].alignment = Alignment(vertical="top", wrap_text=True)
        ws[f"B{row}"].border = border
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)

    for row in (11, 18, 21, 28):
        ws.row_dimensions[row].height = 62

    photo_fields = [
        (7, 10, "상품 대표사진"),
        (14, 17, "사례 첨부사진"),
        (24, 27, "수리방법 사진"),
    ]

    for start_row, end_row, label in photo_fields:
        ws[f"A{start_row}"] = label
        ws[f"A{start_row}"].font = label_font
        ws[f"A{start_row}"].fill = label_fill
        ws[f"A{start_row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"A{start_row}"].border = border
        ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
        ws.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=4)
        photo_cell = ws[f"B{start_row}"]
        photo_cell.value = "이 영역에 사진을 삽입하세요"
        photo_cell.font = Font(size=11, color="64748B")
        photo_cell.fill = photo_fill
        photo_cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in range(start_row, end_row + 1):
            ws.row_dimensions[row].height = 34
            for col in range(1, 5):
                ws.cell(row=row, column=col).border = border
                if col >= 2:
                    ws.cell(row=row, column=col).fill = photo_fill

    validation = DataValidation(
        type="list",
        formula1=f'"{",".join(CASE_CATEGORIES)}"',
        allow_blank=False,
    )
    ws.add_data_validation(validation)
    validation.add(ws["B4"])

    ws.freeze_panes = "A3"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def get_excel_image_bytes_by_row(ws):
    image_map = {}

    for image in getattr(ws, "_images", []):
        try:
            row = image.anchor._from.row + 1
            image_map[row] = image._data()
        except Exception:
            continue

    return image_map


def get_image_in_row_range(image_map, start_row, end_row):
    for row, image_data in image_map.items():
        if start_row <= row <= end_row:
            return image_data
    return None


def read_case_excel_form(uploaded_file):
    wb = load_workbook(BytesIO(uploaded_file.getvalue()))
    ws = wb["사례등록"] if "사례등록" in wb.sheetnames else wb.active
    image_map = get_excel_image_bytes_by_row(ws)

    form_data = {
        "case_date": parse_case_date(ws[CASE_EXCEL_FIELD_CELLS["case_date"]].value),
        "category": normalize_excel_text(ws[CASE_EXCEL_FIELD_CELLS["category"]].value) or "기타",
        "product": normalize_excel_text(ws[CASE_EXCEL_FIELD_CELLS["product"]].value),
        "barcode": normalize_excel_text(ws[CASE_EXCEL_FIELD_CELLS["barcode"]].value),
        "cause": normalize_excel_text(ws[CASE_EXCEL_FIELD_CELLS["cause"]].value),
        "action": normalize_excel_text(ws[CASE_EXCEL_FIELD_CELLS["action"]].value),
        "repair_method": normalize_excel_text(ws[CASE_EXCEL_FIELD_CELLS["repair_method"]].value),
        "prevention": normalize_excel_text(ws[CASE_EXCEL_FIELD_CELLS["prevention"]].value),
    }
    if form_data["category"] not in CASE_CATEGORIES:
        form_data["category"] = "기타"

    for key, row_range in CASE_EXCEL_IMAGE_RANGES.items():
        form_data[key] = get_image_in_row_range(image_map, *row_range)

    return form_data


def insert_case_from_excel_form(uploaded_file):
    form_data = read_case_excel_form(uploaded_file)
    new_case_id = get_next_case_id(form_data["case_date"])

    c.execute("""
    INSERT INTO cases
    (
        case_id,
        category,
        barcode,
        product,
        cause,
        action,
        repair_method,
        prevention,
        product_image,
        case_image,
        case_image_original,
        repair_image,
        repair_image_original
    )
    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
    """,
    (
        new_case_id,
        form_data["category"],
        form_data["barcode"],
        form_data["product"],
        form_data["cause"],
        form_data["action"],
        form_data["repair_method"],
        form_data["prevention"],
        form_data["product_image"],
        form_data["case_image"],
        form_data["case_image"],
        form_data["repair_image"],
        form_data["repair_image"],
    ))

    conn.commit()
    return new_case_id


def image_data_fingerprint(image_data):
    return hashlib.md5(image_data).hexdigest()[:10]


def get_marker_canvas_key(key_prefix, marker_type, image_data):
    reset_key = f"{key_prefix}_{marker_type}_reset_version"
    reset_version = st.session_state.get(reset_key, 0)
    session_version = st.session_state.get("marker_canvas_session_version", 0)
    return (
        f"{key_prefix}_{marker_type}_canvas_"
        f"v{MARK_CANVAS_RENDER_VERSION}_"
        f"s{session_version}_"
        f"{image_data_fingerprint(image_data)}_{reset_version}"
    )


def restore_marker_canvas(key_prefix, marker_type):
    reset_key = f"{key_prefix}_{marker_type}_reset_version"
    st.session_state[reset_key] = st.session_state.get(reset_key, 0) + 1


def reset_marker_canvas_on_image_change(key_prefix, marker_type, image_data):
    image_key = f"{key_prefix}_{marker_type}_canvas_image"
    image_fingerprint = image_data_fingerprint(image_data)

    if st.session_state.get(image_key) != image_fingerprint:
        st.session_state[image_key] = image_fingerprint
        restore_marker_canvas(key_prefix, marker_type)


def clear_marker_session_state(key_prefix, include_canvas_state=False):
    state_suffixes = [
        "_red_restore_image",
        "_red_cleaned_image",
        "_red_cleaned_base",
        "_red_reset_version",
        "_red_canvas_image",
        "_blue_restored_image",
        "_blue_restored_base",
        "_blue_restore_image",
        "_blue_reset_version",
        "_blue_canvas_image",
        "_current_fingerprint",
    ]

    for suffix in state_suffixes:
        st.session_state.pop(f"{key_prefix}{suffix}", None)

    if include_canvas_state:
        canvas_prefixes = (
            f"{key_prefix}_red_canvas_",
            f"{key_prefix}_blue_canvas_",
        )
        for state_key in list(st.session_state.keys()):
            if state_key.startswith(canvas_prefixes):
                st.session_state.pop(state_key, None)


def image_bytes_to_rgb(image_data):
    img = Image.open(BytesIO(image_data))
    img = ImageOps.exif_transpose(img)
    return img.convert("RGB")


def is_readable_image_data(image_data):
    if not image_data:
        return False

    try:
        Image.open(BytesIO(image_data)).verify()
        return True
    except Exception:
        return False


def first_readable_image_data(*image_data_items):
    for image_data in image_data_items:
        if is_readable_image_data(image_data):
            return image_data

    return None


def image_to_jpeg_bytes(img):
    output = BytesIO()
    img.save(output, format="JPEG", quality=92)
    return output.getvalue()


def get_pdf_font(size, bold=False):
    font_candidates = [
        Path("C:/Windows/Fonts/malgunbd.ttf") if bold else Path("C:/Windows/Fonts/malgun.ttf"),
        Path("C:/Windows/Fonts/NanumGothicBold.ttf") if bold else Path("C:/Windows/Fonts/NanumGothic.ttf"),
        Path("C:/Windows/Fonts/gulim.ttc"),
        Path("/usr/share/fonts/truetype/nanum/NanumGothicBold.ttf")
        if bold
        else Path("/usr/share/fonts/truetype/nanum/NanumGothic.ttf"),
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf")
        if bold
        else Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
    ]

    for font_path in font_candidates:
        try:
            if font_path.exists():
                return ImageFont.truetype(str(font_path), size)
        except Exception:
            continue

    return ImageFont.load_default()


def draw_pdf_text(draw, xy, text, font, fill=(15, 23, 42)):
    try:
        draw.text(xy, text, font=font, fill=fill)
    except UnicodeEncodeError:
        draw.text(xy, text.encode("ascii", "replace").decode("ascii"), font=font, fill=fill)


def text_width(draw, text, font):
    try:
        return draw.textbbox((0, 0), text, font=font)[2]
    except UnicodeEncodeError:
        return draw.textbbox((0, 0), text.encode("ascii", "replace").decode("ascii"), font=font)[2]


def wrap_pdf_text(draw, text, font, max_width):
    text = normalize_excel_text(text) or "-"
    lines = []

    for paragraph in text.splitlines() or ["-"]:
        current = ""
        words = paragraph.split(" ")

        for word in words:
            candidate = word if not current else f"{current} {word}"
            if text_width(draw, candidate, font) <= max_width:
                current = candidate
                continue

            if current:
                lines.append(current)
                current = ""

            chunk = ""
            for char in word:
                candidate = f"{chunk}{char}"
                if text_width(draw, candidate, font) <= max_width:
                    chunk = candidate
                else:
                    if chunk:
                        lines.append(chunk)
                    chunk = char
            current = chunk

        lines.append(current or "-")

    return lines


def draw_pdf_wrapped_text(draw, box, text, font, fill=(15, 23, 42), line_gap=8, max_lines=None):
    x, y, width, height = box
    line_height = getattr(font, "size", 18) + line_gap
    lines = wrap_pdf_text(draw, text, font, width)

    if max_lines is not None and len(lines) > max_lines:
        lines = lines[:max_lines]
        lines[-1] = f"{lines[-1][:max(0, len(lines[-1]) - 2)]}..."

    max_fit_lines = max(1, int(height / line_height))
    if len(lines) > max_fit_lines:
        lines = lines[:max_fit_lines]
        lines[-1] = f"{lines[-1][:max(0, len(lines[-1]) - 2)]}..."

    for line in lines:
        draw_pdf_text(draw, (x, y), line, font, fill)
        y += line_height


def draw_pdf_fitted_wrapped_text(draw, box, text, font, fill=(15, 23, 42), line_gap=5, min_size=12):
    x, y, width, height = box
    fitted_font = font

    for size in range(getattr(font, "size", 18), min_size - 1, -1):
        try:
            candidate_font = font.font_variant(size=size)
        except Exception:
            candidate_font = font

        line_height = getattr(candidate_font, "size", size) + line_gap
        lines = wrap_pdf_text(draw, text, candidate_font, width)

        if len(lines) * line_height <= height:
            fitted_font = candidate_font
            break
    else:
        line_height = getattr(fitted_font, "size", 18) + line_gap
        lines = wrap_pdf_text(draw, text, fitted_font, width)

    line_height = getattr(fitted_font, "size", 18) + line_gap
    max_fit_lines = max(1, int(height / line_height))
    if len(lines) > max_fit_lines:
        lines = lines[:max_fit_lines]

    for line in lines:
        draw_pdf_text(draw, (x, y), line, fitted_font, fill)
        y += line_height


def draw_pdf_image(page, image_data, box):
    if not image_data:
        return

    try:
        img = Image.open(BytesIO(image_data))
        img = ImageOps.exif_transpose(img).convert("RGB")
    except Exception:
        return

    x, y, width, height = box
    img.thumbnail((width, height), Image.LANCZOS)
    paste_x = x + (width - img.width) // 2
    paste_y = y + (height - img.height) // 2
    page.paste(img, (paste_x, paste_y))


def create_detail_pdf(data):
    page_size = (1240, 1754)
    margin = 72
    pages = []

    def new_page():
        page = Image.new("RGB", page_size, "white")
        pages.append(page)
        return page, ImageDraw.Draw(page)

    page, draw = new_page()
    title_font = get_pdf_font(34, bold=True)
    label_font = get_pdf_font(20, bold=True)
    value_font = get_pdf_font(23, bold=True)
    body_font = get_pdf_font(21)
    small_font = get_pdf_font(17)

    black = (15, 23, 42)
    gray = (100, 116, 139)
    border = (203, 213, 225)
    accent_colors = [(56, 189, 248), (34, 197, 94), (245, 158, 11), (244, 114, 182)]

    y = 58
    draw_pdf_text(draw, (margin, y), "사례 상세보기", title_font, black)
    y += 62

    info_x = margin
    info_y = y
    info_w = page_size[0] - margin * 2
    info_h = 416
    draw.rounded_rectangle((info_x, info_y, info_x + info_w, info_y + info_h), radius=14, outline=border, width=2)

    image_box = (info_x + 24, info_y + 24, 260, 368)
    draw.rounded_rectangle(
        (image_box[0], image_box[1], image_box[0] + image_box[2], image_box[1] + image_box[3]),
        radius=10,
        outline=border,
        width=2,
    )
    draw_pdf_image(page, data[8], image_box)

    meta_x = info_x + 320
    meta_y = info_y + 24
    cell_gap = 18
    half_cell_w = (info_w - 360 - cell_gap) // 2
    full_cell_w = info_w - 360
    meta_items = [
        ("사례번호", data[0], 0, 0, half_cell_w, 88),
        ("유형", data[1], 0, 1, half_cell_w, 88),
        ("상품명", data[3], 1, 0, full_cell_w, 154),
        ("바코드", data[2], 2, 0, full_cell_w, 88),
    ]

    for label, value, row, col, cell_w, cell_h in meta_items:
        x = meta_x + col * (half_cell_w + cell_gap)
        cell_y = meta_y + [0, 106, 278][row]
        draw.rounded_rectangle((x, cell_y, x + cell_w, cell_y + cell_h), radius=10, outline=border, width=2)
        draw_pdf_text(draw, (x + 18, cell_y + 16), label, label_font, gray)
        draw_pdf_fitted_wrapped_text(
            draw,
            (x + 18, cell_y + 50, cell_w - 36, cell_h - 62),
            value,
            value_font,
            black,
            line_gap=4,
            min_size=13,
        )

    y += info_h + 32

    section_items = [
        ("원인", data[4]),
        ("조치방법", data[5]),
        ("수리방법", data[6]),
        ("방지대책", data[7]),
    ]
    section_w = (page_size[0] - margin * 2 - 24) // 2
    section_h = 190

    for idx, (title, body) in enumerate(section_items):
        row = idx // 2
        col = idx % 2
        x = margin + col * (section_w + 24)
        section_y = y + row * (section_h + 24)
        draw.rounded_rectangle((x, section_y, x + section_w, section_y + section_h), radius=12, outline=border, width=2)
        draw.line((x + 1, section_y + 12, x + 1, section_y + section_h - 12), fill=accent_colors[idx], width=7)
        draw_pdf_text(draw, (x + 24, section_y + 20), title, label_font, black)
        draw_pdf_wrapped_text(
            draw,
            (x + 24, section_y + 62, section_w - 48, section_h - 84),
            body,
            body_font,
            black,
            line_gap=8,
        )

    y += (section_h * 2) + 72

    photo_items = [("사례 사진", data[9]), ("수리 사진", data[10])]
    photo_w = (page_size[0] - margin * 2 - 24) // 2
    photo_h = 420

    if y + photo_h + 72 > page_size[1]:
        page, draw = new_page()
        y = 58

    for idx, (title, image_data) in enumerate(photo_items):
        x = margin + idx * (photo_w + 24)
        draw_pdf_text(draw, (x, y), title, label_font, black)
        box = (x, y + 38, photo_w, photo_h)
        draw.rounded_rectangle((box[0], box[1], box[0] + box[2], box[1] + box[3]), radius=12, outline=border, width=2)
        draw_pdf_image(page, image_data, (box[0] + 12, box[1] + 12, box[2] - 24, box[3] - 24))

    draw_pdf_text(draw, (margin, page_size[1] - 58), f"사례번호 {data[0]}", small_font, gray)

    buffer = BytesIO()
    pages[0].save(buffer, format="PDF", resolution=150.0, save_all=True, append_images=pages[1:])
    buffer.seek(0)
    return buffer.getvalue()


def resize_for_marker_canvas(img):
    img = ImageOps.exif_transpose(img).convert("RGB")

    if img.width <= MARK_CANVAS_MAX_WIDTH:
        return img.copy()

    ratio = MARK_CANVAS_MAX_WIDTH / img.width
    height = max(1, int(img.height * ratio))
    return img.resize((MARK_CANVAS_MAX_WIDTH, height), Image.LANCZOS)


def canvas_rgba_to_jpeg_bytes(image_data, background_img=None):
    canvas_img = Image.fromarray(image_data.astype("uint8"), "RGBA")

    if background_img is None:
        background = Image.new("RGBA", canvas_img.size, (0, 0, 0, 0))
    else:
        background = background_img.convert("RGBA").resize(canvas_img.size)

    background.alpha_composite(canvas_img)
    return image_to_jpeg_bytes(background.convert("RGB"))


def canvas_objects_to_original_jpeg_bytes(original_img, canvas_img, canvas_objects, fallback_image_data=None):
    marked_img = ImageOps.exif_transpose(original_img).convert("RGB").copy()
    canvas_w = max(1, canvas_img.width)
    canvas_h = max(1, canvas_img.height)
    scale_x = marked_img.width / canvas_w
    scale_y = marked_img.height / canvas_h
    draw = ImageDraw.Draw(marked_img)
    drew_object = False

    for obj in canvas_objects or []:
        obj_type = str(obj.get("type", "")).lower()
        if obj_type not in {"circle", "ellipse"}:
            continue

        left = float(obj.get("left") or 0)
        top = float(obj.get("top") or 0)
        width = float(obj.get("width") or (float(obj.get("radius") or 0) * 2) or 0)
        height = float(obj.get("height") or (float(obj.get("radius") or 0) * 2) or width)
        obj_scale_x = float(obj.get("scaleX") or 1)
        obj_scale_y = float(obj.get("scaleY") or 1)
        stroke = obj.get("stroke") or "#ff2d2d"
        stroke_width = max(
            1,
            int(round(float(obj.get("strokeWidth") or 1) * ((scale_x + scale_y) / 2))),
        )

        x1 = int(round(left * scale_x))
        y1 = int(round(top * scale_y))
        x2 = int(round((left + width * obj_scale_x) * scale_x))
        y2 = int(round((top + height * obj_scale_y) * scale_y))
        if x2 <= x1 or y2 <= y1:
            continue

        draw.ellipse((x1, y1, x2, y2), outline=stroke, width=stroke_width)
        drew_object = True

    if drew_object:
        return image_to_jpeg_bytes(marked_img)

    if fallback_image_data is not None:
        return canvas_rgba_to_original_jpeg_bytes(fallback_image_data, original_img)

    return image_to_jpeg_bytes(marked_img)


def canvas_rgba_to_original_jpeg_bytes(image_data, original_img):
    canvas_img = Image.fromarray(image_data.astype("uint8"), "RGBA").convert("RGB")
    original_img = ImageOps.exif_transpose(original_img).convert("RGB")
    resized = canvas_img.resize(original_img.size, Image.LANCZOS)
    return image_to_jpeg_bytes(resized)


def image_to_png_data_url(img):
    output = BytesIO()
    img = ImageOps.exif_transpose(img).convert("RGB")
    img.save(output, format="PNG")
    image_base64 = base64.b64encode(output.getvalue()).decode("ascii")
    return f"data:image/png;base64,{image_base64}"


def get_canvas_initial_drawing(background_img):
    return {
        "version": "4.4.0",
        "objects": [],
        "background": "",
        "backgroundImage": {
            "type": "image",
            "version": "4.4.0",
            "originX": "left",
            "originY": "top",
            "left": 0,
            "top": 0,
            "width": background_img.width,
            "height": background_img.height,
            "scaleX": 1,
            "scaleY": 1,
            "angle": 0,
            "flipX": False,
            "flipY": False,
            "opacity": 1,
            "src": image_to_png_data_url(background_img),
            "crossOrigin": None,
            "filters": [],
        },
    }

def draw_part_circle(img, x_percent, y_percent, radius_percent, color, width):
    marked_img = img.copy()
    draw = ImageDraw.Draw(marked_img)

    x = int(marked_img.width * x_percent / 100)
    y = int(marked_img.height * y_percent / 100)
    radius = int(min(marked_img.size) * radius_percent / 100)

    draw.ellipse(
        (x - radius, y - radius, x + radius, y + radius),
        outline=color,
        width=width,
    )

    return marked_img


def part_marker_tool(
    image_data,
    key_prefix,
    title="부위 표시",
    restore_image_data=None,
):
    if not image_data:
        return None

    if restore_image_data is None:
        restore_image_data = image_data

    try:
        img = image_bytes_to_rgb(image_data)
    except Exception:
        st.warning("이미지를 불러올 수 없습니다.")
        return image_data

    with st.expander(title, expanded=False):
        enabled = st.checkbox(
            "사진에 동그라미 표시",
            key=f"{key_prefix}_enabled",
        )

        if not enabled:
            st.image(img, caption="원본 사진", width=MARK_CANVAS_MAX_WIDTH)
            return image_data

        if st_canvas is not None:
            canvas_img = resize_for_marker_canvas(img)
            canvas_key = get_marker_canvas_key(key_prefix, "part", image_data)

            control_col1, control_col2 = st.columns(2)

            with control_col1:
                color_name = st.selectbox(
                    "표시 색상",
                    list(MARK_COLORS.keys()),
                    key=f"{key_prefix}_canvas_color",
                )

            with control_col2:
                line_width = st.slider(
                    "선 굵기",
                    2,
                    18,
                    8,
                    key=f"{key_prefix}_canvas_width",
                )

            canvas_result = st_canvas(
                fill_color="rgba(255, 255, 255, 0)",
                stroke_width=line_width,
                stroke_color=MARK_COLORS[color_name],
                initial_drawing=get_canvas_initial_drawing(canvas_img),
                update_streamlit=True,
                height=canvas_img.height,
                width=canvas_img.width,
                drawing_mode="circle",
                key=canvas_key,
            )

            canvas_objects = (
                canvas_result.json_data.get("objects", [])
                if canvas_result.json_data
                else []
            )

            if canvas_result.image_data is not None and canvas_objects:
                return canvas_objects_to_original_jpeg_bytes(
                    img,
                    canvas_img,
                    canvas_objects,
                    canvas_result.image_data,
                )

            return image_data

        st.info(
            "마우스로 직접 표시하려면 streamlit-drawable-canvas 설치가 필요합니다. "
            "설치 전에는 아래 위치 조절 방식으로 저장됩니다."
        )

        pos_col1, pos_col2 = st.columns(2)

        with pos_col1:
            x_percent = st.slider(
                "가로 위치",
                0,
                100,
                50,
                key=f"{key_prefix}_x",
            )
            radius_percent = st.slider(
                "원 크기",
                3,
                45,
                18,
                key=f"{key_prefix}_radius",
            )

        with pos_col2:
            y_percent = st.slider(
                "세로 위치",
                0,
                100,
                50,
                key=f"{key_prefix}_y",
            )
            line_width = st.slider(
                "선 굵기",
                2,
                18,
                8,
                key=f"{key_prefix}_width",
            )

        color_name = st.selectbox(
            "표시 색상",
            list(MARK_COLORS.keys()),
            key=f"{key_prefix}_color",
        )

        marked_img = draw_part_circle(
            img,
            x_percent,
            y_percent,
            radius_percent,
            MARK_COLORS[color_name],
            line_width,
        )

        st.image(marked_img, caption="저장될 표시 사진", width=MARK_CANVAS_MAX_WIDTH)
        return image_to_jpeg_bytes(marked_img)


def edit_red_circle_marker_tool(
    image_data,
    key_prefix,
    title=None,
    restore_image_data=None,
):
    if not image_data:
        return None

    restore_state_key = f"{key_prefix}_red_restore_image"
    if restore_image_data is None and restore_state_key in st.session_state:
        restore_image_data = st.session_state[restore_state_key]

    if restore_image_data is None and key_prefix.startswith("edit_case_marker_"):
        restore_case_id = key_prefix[len("edit_case_marker_"):]
        restore_image_data = get_case_image_original(restore_case_id)

    original_image_fingerprint = image_data_fingerprint(image_data)
    cleaned_image_key = f"{key_prefix}_red_cleaned_image"
    cleaned_base_key = f"{key_prefix}_red_cleaned_base"

    if (
        st.session_state.get(cleaned_base_key) == original_image_fingerprint
        and cleaned_image_key in st.session_state
    ):
        image_data = st.session_state[cleaned_image_key]

    try:
        img = image_bytes_to_rgb(image_data)
    except Exception:
        st.warning("이미지를 불러올 수 없습니다.")
        return image_data

    with st.expander("사례사진 빨간 원 표시", expanded=True):
        enabled = st.checkbox(
            "기존 사진에 빨간 원 표시",
            value=True,
            key=f"{key_prefix}_red_enabled",
        )

        if not enabled:
            st.image(img, caption="원본 사진", width=MARK_CANVAS_MAX_WIDTH)
            return image_data

        if st_canvas is not None:
            canvas_img = resize_for_marker_canvas(img)
            canvas_key = get_marker_canvas_key(key_prefix, "red", image_data)
            restore_col, _ = st.columns([0.7, 9.3], gap="small")

            if restore_col.button(
                "사진 복원",
                key=f"{key_prefix}_red_remove_saved_btn",
                use_container_width=True,
            ):
                if not restore_image_data:
                    st.info("복원할 원본 사진이 없습니다.")
                else:
                    image_data = restore_image_data
                    img = image_bytes_to_rgb(image_data)
                    canvas_img = resize_for_marker_canvas(img)
                    st.session_state[cleaned_image_key] = restore_image_data
                    st.session_state[cleaned_base_key] = original_image_fingerprint
                    reset_key = f"{key_prefix}_red_reset_version"
                    st.session_state[reset_key] = (
                        st.session_state.get(reset_key, 0) + 1
                    )
                    canvas_key = get_marker_canvas_key(
                        key_prefix,
                        "red",
                        image_data,
                    )
                    st.success("원본 사진으로 복원했습니다. 수정 저장을 눌러 반영하세요.")

            line_width = st.slider(
                "원 두께",
                2,
                18,
                8,
                key=f"{key_prefix}_red_canvas_width",
            )

            canvas_result = st_canvas(
                fill_color="rgba(255, 255, 255, 0)",
                stroke_width=line_width,
                stroke_color="#ff2d2d",
                initial_drawing=get_canvas_initial_drawing(canvas_img),
                update_streamlit=True,
                height=canvas_img.height,
                width=canvas_img.width,
                drawing_mode="circle",
                key=canvas_key,
            )

            canvas_objects = (
                canvas_result.json_data.get("objects", [])
                if canvas_result.json_data
                else []
            )

            if canvas_result.image_data is not None and canvas_objects:
                return canvas_objects_to_original_jpeg_bytes(
                    img,
                    canvas_img,
                    canvas_objects,
                    canvas_result.image_data,
                )

            return image_data

        st.info(
            "마우스 표시 도구를 사용할 수 없어 위치 조절 방식으로 저장됩니다."
        )

        pos_col1, pos_col2 = st.columns(2)

        with pos_col1:
            x_percent = st.slider(
                "가로 위치",
                0,
                100,
                50,
                key=f"{key_prefix}_red_x",
            )
            radius_percent = st.slider(
                "원 크기",
                3,
                45,
                18,
                key=f"{key_prefix}_red_radius",
            )

        with pos_col2:
            y_percent = st.slider(
                "세로 위치",
                0,
                100,
                50,
                key=f"{key_prefix}_red_y",
            )
            line_width = st.slider(
                "원 두께",
                2,
                18,
                8,
                key=f"{key_prefix}_red_width",
            )

        marked_img = draw_part_circle(
            img,
            x_percent,
            y_percent,
            radius_percent,
            "#ff2d2d",
            line_width,
        )

        st.image(marked_img, caption="저장될 표시 사진", width=MARK_CANVAS_MAX_WIDTH)
        return image_to_jpeg_bytes(marked_img)


def blue_circle_marker_tool(
    image_data,
    key_prefix,
    title=None,
    restore_image_data=None,
    show_photo_buttons=True,
):
    if not image_data:
        return None

    restore_state_key = f"{key_prefix}_blue_restore_image"
    if restore_image_data is None and restore_state_key in st.session_state:
        restore_image_data = st.session_state[restore_state_key]

    if restore_image_data is None and key_prefix.startswith("edit_repair_marker_"):
        restore_case_id = key_prefix[len("edit_repair_marker_"):]
        restore_image_data = get_repair_image_original(restore_case_id)

    if restore_image_data is None:
        restore_image_data = image_data

    base_image_fingerprint = image_data_fingerprint(image_data)
    restored_image_key = f"{key_prefix}_blue_restored_image"
    restored_base_key = f"{key_prefix}_blue_restored_base"

    if (
        st.session_state.get(restored_base_key) == base_image_fingerprint
        and restored_image_key in st.session_state
    ):
        image_data = st.session_state[restored_image_key]

    try:
        img = image_bytes_to_rgb(image_data)
    except Exception:
        st.warning("?대?吏瑜?遺덈윭?????놁뒿?덈떎.")
        return image_data

    with st.expander("수리사진 파란 원 표시", expanded=True):
        enabled = st.checkbox(
            "사진에 파란 원 표시",
            value=True,
            key=f"{key_prefix}_blue_enabled",
        )

        if not enabled:
            st.image(img, caption="원본 사진", width=MARK_CANVAS_MAX_WIDTH)
            return image_data

        if st_canvas is not None:
            canvas_img = resize_for_marker_canvas(img)
            canvas_key = get_marker_canvas_key(key_prefix, "blue", image_data)

            line_width = st.slider(
                "원 두께",
                2,
                18,
                8,
                key=f"{key_prefix}_blue_canvas_width",
            )

            restore_col, _ = st.columns([0.7, 9.3], gap="small")

            if show_photo_buttons and restore_col.button(
                "사진 복원",
                key=f"{key_prefix}_blue_restore_btn",
                use_container_width=True,
            ):
                image_data = restore_image_data
                img = image_bytes_to_rgb(image_data)
                canvas_img = resize_for_marker_canvas(img)
                st.session_state[restored_image_key] = restore_image_data
                st.session_state[restored_base_key] = base_image_fingerprint
                restore_marker_canvas(key_prefix, "blue")
                canvas_key = get_marker_canvas_key(key_prefix, "blue", image_data)
                st.success("원본 사진으로 복원했습니다. 수정 저장을 눌러 반영하세요.")

            canvas_result = st_canvas(
                fill_color="rgba(255, 255, 255, 0)",
                stroke_width=line_width,
                stroke_color="#0a84ff",
                initial_drawing=get_canvas_initial_drawing(canvas_img),
                update_streamlit=True,
                height=canvas_img.height,
                width=canvas_img.width,
                drawing_mode="circle",
                key=canvas_key,
            )

            canvas_objects = (
                canvas_result.json_data.get("objects", [])
                if canvas_result.json_data
                else []
            )

            if canvas_result.image_data is not None and canvas_objects:
                return canvas_objects_to_original_jpeg_bytes(
                    img,
                    canvas_img,
                    canvas_objects,
                    canvas_result.image_data,
                )

            return image_data

        st.info(
            "마우스 표시 도구를 사용할 수 없어 파란 원 표시를 할 수 없습니다."
        )

        st.image(img, caption="원본 사진", width=MARK_CANVAS_MAX_WIDTH)
        return image_data


def detail_html(value):
    text = str(value).strip() if value else "-"
    return escape(text).replace("\n", "<br>")


def get_next_case_id(case_date, current_case_id=None):
    case_date_key = case_date.strftime("%Y%m%d")

    if current_case_id and current_case_id.startswith(f"{case_date_key}-"):
        return current_case_id

    c.execute(
        """
        SELECT COALESCE(MAX(CAST(substr(case_id, 10) AS INTEGER)), 0)
        FROM cases
        WHERE case_id LIKE ?
          AND case_id != ?
        """,
        (f"{case_date_key}-%", current_case_id or "")
    )

    count = c.fetchone()[0] + 1

    return f"{case_date_key}-{count:03d}"

# ==========================
# DB 연결

def render_return_case_system():
    global conn, c

    # ==========================

    if not DB_PATH.exists():
        st.error(f"기존 반품/AS DB 파일을 찾을 수 없습니다: {DB_PATH}")
        return

    conn = sqlite3.connect(str(DB_PATH), check_same_thread=False)
    c = conn.cursor()

    c.execute("""
    CREATE TABLE IF NOT EXISTS cases(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        case_id TEXT,
        category TEXT,
        barcode TEXT,
        product TEXT,
        cause TEXT,
        action TEXT,
        repair_method TEXT,
        prevention TEXT,

        product_image BLOB,
        case_image BLOB,
        case_image_original BLOB,
        repair_image BLOB,
        repair_image_original BLOB
    )
    """)

    conn.commit()

    c.execute("PRAGMA table_info(cases)")
    case_columns = {row[1] for row in c.fetchall()}

    if "case_image_original" not in case_columns:
        c.execute("ALTER TABLE cases ADD COLUMN case_image_original BLOB")
        c.execute("""
        UPDATE cases
        SET case_image_original = case_image
        WHERE case_image_original IS NULL
        """)
        conn.commit()

    if "repair_image_original" not in case_columns:
        c.execute("ALTER TABLE cases ADD COLUMN repair_image_original BLOB")
        c.execute("""
        UPDATE cases
        SET repair_image_original = repair_image
        WHERE repair_image_original IS NULL
        """)
        conn.commit()

    c.execute("UPDATE cases SET category='누락' WHERE category='변심'")
    conn.commit()

    def get_case_image_original(case_id):
        c.execute(
            "SELECT case_image_original FROM cases WHERE case_id = ?",
            (case_id,),
        )
        row = c.fetchone()
        return row[0] if row and row[0] else None


    def get_repair_image_original(case_id):
        c.execute(
            "SELECT repair_image_original FROM cases WHERE case_id = ?",
            (case_id,),
        )
        row = c.fetchone()
        return row[0] if row and row[0] else None

    # ==========================
    # 화면 스타일
    # ==========================

    st.markdown("""
    <style>

    .stApp{
        background:
        radial-gradient(circle at 18% 0%, rgba(31, 171, 150, 0.26), transparent 32%),
        radial-gradient(circle at 84% 38%, rgba(18, 104, 94, 0.30), transparent 34%),
        linear-gradient(135deg, #031b18 0%, #052421 45%, #073a34 100%);
    }
    html,
    body,
    [data-testid="stAppViewContainer"],
    .stApp {
        margin:0 !important;
        padding:0 !important;
        min-height:100vh !important;
    }

    #MainMenu,
    footer,
    header[data-testid="stHeader"],
    div[data-testid="stToolbar"],
    div[data-testid="stDecoration"],
    div[data-testid="stStatusWidget"] {
        display:none !important;
        visibility:hidden !important;
    }

    /* 전체 글씨 */
    html,body,
    h1,h2,h3,h4,h5,h6,
    p,label,span,div{
        color:white !important;
    }
    [data-testid="column"]:first-child{
        border-right:1px solid rgba(255,255,255,0.08);
        padding-right:20px;
    }

    /* 입력창 */
    .stTextInput input{
        background:#052421 !important;
        color:white !important;
        border:none !important;
        border-radius:10px !important;
    }

    div[class*="st-key-search_keyword"] div[data-baseweb="input"] {
        background: rgba(5, 36, 33, 0.96) !important;
        border: 1px solid rgba(40, 217, 197, 0.58) !important;
        border-radius: 12px !important;
        box-shadow:
            inset 0 0 0 1px rgba(255,255,255,0.05),
            0 0 0 1px rgba(12, 145, 132, 0.14),
            0 8px 18px rgba(0,0,0,0.18) !important;
    }

    div[class*="st-key-search_keyword"] div[data-baseweb="input"]:focus-within {
        border-color: rgba(66, 245, 221, 0.92) !important;
        box-shadow:
            inset 0 0 0 1px rgba(255,255,255,0.08),
            0 0 0 3px rgba(28, 203, 184, 0.18),
            0 10px 22px rgba(0,0,0,0.22) !important;
    }

    div[class*="st-key-search_keyword"] input {
        height: 42px !important;
        background: transparent !important;
        border: none !important;
        padding: 0 13px !important;
        font-weight: 750 !important;
    }

    /* 텍스트영역 */
    .stTextArea textarea{
        background:#052421 !important;
        color:white !important;
        border:none !important;
        border-radius:10px !important;
    }

    /* 파일 업로더 */
    div[data-testid="stFileUploader"] section,
    div[data-testid="stFileUploaderDropzone"] {
        background:#052421 !important;
        border:1px dashed rgba(255,255,255,0.24) !important;
        border-radius:10px !important;
    }

    div[data-testid="stFileUploader"] section *,
    div[data-testid="stFileUploaderDropzone"] *,
    div[data-testid="stFileUploaderFile"] * {
        color:white !important;
    }

    div[data-testid="stFileUploader"] section button,
    div[data-testid="stFileUploaderDropzone"] button {
        background:rgba(255,255,255,0.10) !important;
        color:white !important;
        border:1px solid rgba(255,255,255,0.18) !important;
        border-radius:8px !important;
    }

    div[data-testid="stFileUploader"] section button:hover,
    div[data-testid="stFileUploaderDropzone"] button:hover {
        background:rgba(255,255,255,0.18) !important;
        border-color:rgba(255,255,255,0.30) !important;
    }

    div[data-testid="stFileUploaderFile"] {
        background:#111827 !important;
        border:1px solid rgba(255,255,255,0.14) !important;
        border-radius:10px !important;
    }

    .st-key-edit_prevention {
        margin-top:0.85rem !important;
    }

    .edit-prevention-top-gap {
        height:0.45rem;
        line-height:0;
    }

    .edit-button-top-gap {
        height:0.95rem;
        line-height:0;
    }

    .excel-download-gap {
        height:0;
        line-height:0;
    }

    /* 셀렉트 */
    .stSelectbox div[data-baseweb="select"],
    .stSelectbox div[data-baseweb="select"] > div,
    .stSelectbox div[role="combobox"],
    div[data-baseweb="select"],
    div[data-baseweb="select"] > div {
        background:#052421 !important;
        color:white !important;
        border:none !important;
        border-radius:10px !important;
    }

    .stSelectbox div[data-baseweb="select"] *,
    .stSelectbox div[role="combobox"] *,
    div[data-baseweb="select"] *,
    .stDateInput div[data-baseweb="input"] *,
    .stDateInput input {
        color:white !important;
    }

    .stSelectbox svg,
    div[data-baseweb="select"] svg {
        color:white !important;
        fill:white !important;
    }

    /* 검색기준 라디오 */
    div[data-testid="stRadio"] > label,
    div[data-testid="stRadio"] > div > label {
        background:transparent !important;
        border:none !important;
        padding:0 !important;
        margin-bottom:4px !important;
    }

    div[data-testid="stRadio"] div[role="radiogroup"] {
        display:flex !important;
        flex-wrap:wrap !important;
        justify-content:flex-start !important;
        gap:5px 8px !important;
        overflow:hidden !important;
        padding-bottom:0 !important;
        max-width:100% !important;
    }

    div[data-testid="stRadio"] div[role="radiogroup"] label {
        flex:0 1 auto !important;
        min-width:auto !important;
        min-height:21px !important;
        margin:0 !important;
        background:transparent !important;
        border:none !important;
        border-radius:0 !important;
        padding:0 1px !important;
        box-shadow:none !important;
        white-space:nowrap !important;
        display:flex !important;
        align-items:center !important;
        justify-content:flex-start !important;
    }

    div[data-testid="stRadio"] div[role="radiogroup"] label:hover {
        background:transparent !important;
        border-color:transparent !important;
    }

    div[data-testid="stRadio"] div[role="radiogroup"] label:has(input:checked) {
        background:transparent !important;
        border-color:transparent !important;
    }

    div[data-testid="stRadio"] div[role="radiogroup"] label *,
    div[data-testid="stRadio"] div[role="radiogroup"] * {
        color:white !important;
        font-size:13px !important;
        font-weight:750 !important;
        white-space:nowrap !important;
        align-items:center !important;
        justify-content:center !important;
        text-align:center !important;
    }

    div[data-testid="stRadio"] div[role="radiogroup"] label:has(input:checked) * {
        color:#5eead4 !important;
    }

    div[data-testid="stRadio"] input {
        accent-color:#0d9488 !important;
        width:11px !important;
        height:11px !important;
    }

    @media (max-width: 1200px) {
        div[data-testid="stRadio"] div[role="radiogroup"] {
            gap:4px 6px !important;
        }

        div[data-testid="stRadio"] div[role="radiogroup"] label *,
        div[data-testid="stRadio"] div[role="radiogroup"] * {
            font-size:12px !important;
        }

        div[data-testid="stRadio"] input {
            width:10px !important;
            height:10px !important;
        }
    }

    @media (max-width: 980px) {
        div[data-testid="stRadio"] div[role="radiogroup"] {
            gap:3px 5px !important;
        }

        div[data-testid="stRadio"] div[role="radiogroup"] label *,
        div[data-testid="stRadio"] div[role="radiogroup"] * {
            font-size:11px !important;
        }
    }

    .stDateInput div[data-baseweb="input"],
    .stDateInput input {
        background:#052421 !important;
        border:none !important;
        border-radius:10px !important;
    }

    .stTextInput input::placeholder,
    .stTextArea textarea::placeholder,
    .stDateInput input::placeholder {
        color:rgba(255,255,255,0.55) !important;
    }

    div[data-baseweb="popover"],
    div[data-baseweb="popover"] > div,
    div[data-baseweb="menu"],
    div[data-baseweb="select-dropdown"],
    div[data-baseweb="select-dropdown"] *,
    ul[role="listbox"] {
        background:#052421 !important;
        color:white !important;
        border:1px solid rgba(255,255,255,0.16) !important;
        border-radius:10px !important;
        box-shadow:0 12px 28px rgba(0,0,0,0.32) !important;
    }

    div[data-baseweb="popover"] *,
    div[data-baseweb="menu"] *,
    div[data-baseweb="select-dropdown"] *,
    ul[role="listbox"] *,
    li[role="option"],
    div[role="option"] {
        color:white !important;
    }

    li[role="option"],
    div[role="option"] {
        background:#052421 !important;
    }

    li[role="option"] *,
    div[role="option"] * {
        background:transparent !important;
        color:white !important;
    }

    li[role="option"]:hover,
    div[role="option"]:hover {
        background:#1f2937 !important;
        color:white !important;
    }

    li[aria-selected="true"],
    div[aria-selected="true"] {
        background:#1f2937 !important;
        box-shadow:
            inset 4px 0 0 rgba(255,255,255,0.82),
            inset 0 0 0 1px rgba(255,255,255,0.16) !important;
        color:white !important;
    }

    ul[role="listbox"] li[role="option"]:hover,
    ul[role="listbox"] div[role="option"]:hover,
    ul[role="listbox"] li[role="option"][aria-selected="true"],
    ul[role="listbox"] div[role="option"][aria-selected="true"],
    ul[role="listbox"] li[role="option"][data-highlighted="true"],
    ul[role="listbox"] div[role="option"][data-highlighted="true"] {
        background:#1f2937 !important;
        color:white !important;
    }

    li[aria-selected="true"] *,
    div[aria-selected="true"] * {
        background:transparent !important;
        color:white !important;
    }

    div[data-baseweb="calendar"],
    div[data-baseweb="calendar"] *,
    div[data-baseweb="datepicker"],
    div[data-baseweb="datepicker"] * {
        background:#111827 !important;
        color:white !important;
    }

    div[data-baseweb="calendar"] button,
    div[data-baseweb="calendar"] [role="button"],
    div[data-baseweb="datepicker"] button,
    div[data-baseweb="datepicker"] [role="button"] {
        background:transparent !important;
        color:white !important;
        border-radius:8px !important;
    }

    div[data-baseweb="calendar"] button:hover,
    div[data-baseweb="calendar"] [role="button"]:hover,
    div[data-baseweb="datepicker"] button:hover,
    div[data-baseweb="datepicker"] [role="button"]:hover {
        background:rgba(255,255,255,0.16) !important;
    }

    div[data-baseweb="calendar"] [aria-selected="true"],
    div[data-baseweb="datepicker"] [aria-selected="true"] {
        background:#111827 !important;
        color:white !important;
    }

    div[data-baseweb="calendar"] button[aria-selected="true"],
    div[data-baseweb="calendar"] [role="button"][aria-selected="true"],
    div[data-baseweb="datepicker"] button[aria-selected="true"],
    div[data-baseweb="datepicker"] [role="button"][aria-selected="true"] {
        background:transparent !important;
        color:white !important;
        border:1px solid rgba(255,255,255,0.82) !important;
        border-radius:999px !important;
    }

    div[data-baseweb="calendar"] [role="grid"],
    div[data-baseweb="calendar"] [role="row"],
    div[data-baseweb="calendar"] [role="gridcell"],
    div[data-baseweb="calendar"] table,
    div[data-baseweb="calendar"] thead,
    div[data-baseweb="calendar"] tbody,
    div[data-baseweb="calendar"] tr,
    div[data-baseweb="calendar"] td,
    div[data-baseweb="calendar"] th,
    div[data-baseweb="datepicker"] [role="grid"],
    div[data-baseweb="datepicker"] [role="row"],
    div[data-baseweb="datepicker"] [role="gridcell"],
    div[data-baseweb="datepicker"] table,
    div[data-baseweb="datepicker"] thead,
    div[data-baseweb="datepicker"] tbody,
    div[data-baseweb="datepicker"] tr,
    div[data-baseweb="datepicker"] td,
    div[data-baseweb="datepicker"] th {
        background:#111827 !important;
    }

    div[data-baseweb="calendar"] [role="gridcell"] *,
    div[data-baseweb="datepicker"] [role="gridcell"] * {
        background:transparent !important;
    }

    div[data-baseweb="calendar"] [role="gridcell"]:empty,
    div[data-baseweb="calendar"] [role="gridcell"] > div:empty,
    div[data-baseweb="calendar"] [aria-disabled="true"],
    div[data-baseweb="calendar"] [disabled],
    div[data-baseweb="datepicker"] [role="gridcell"]:empty,
    div[data-baseweb="datepicker"] [role="gridcell"] > div:empty,
    div[data-baseweb="datepicker"] [aria-disabled="true"],
    div[data-baseweb="datepicker"] [disabled] {
        background:#111827 !important;
        color:rgba(255,255,255,0.32) !important;
        border-radius:0 !important;
    }

    div[data-baseweb="calendar"] [role="gridcell"] [aria-selected="true"],
    div[data-baseweb="datepicker"] [role="gridcell"] [aria-selected="true"],
    div[data-baseweb="calendar"] [role="gridcell"][aria-selected="true"],
    div[data-baseweb="datepicker"] [role="gridcell"][aria-selected="true"] {
        background:#111827 !important;
        color:white !important;
        border-radius:0 !important;
    }

    div[data-baseweb="calendar"] [role="gridcell"] button[aria-selected="true"],
    div[data-baseweb="calendar"] [role="gridcell"] [role="button"][aria-selected="true"],
    div[data-baseweb="datepicker"] [role="gridcell"] button[aria-selected="true"],
    div[data-baseweb="datepicker"] [role="gridcell"] [role="button"][aria-selected="true"] {
        background:transparent !important;
        color:white !important;
        border:1px solid rgba(255,255,255,0.82) !important;
        border-radius:999px !important;
    }

    div[data-baseweb="calendar"] [role="gridcell"][data-selected="true"],
    div[data-baseweb="calendar"] [role="gridcell"][data-highlighted="true"],
    div[data-baseweb="calendar"] [role="gridcell"][data-in-range="true"],
    div[data-baseweb="datepicker"] [role="gridcell"][data-selected="true"],
    div[data-baseweb="datepicker"] [role="gridcell"][data-highlighted="true"],
    div[data-baseweb="datepicker"] [role="gridcell"][data-in-range="true"] {
        background:#111827 !important;
    }

    /* 버튼 */
    .stButton button,
    .stDownloadButton button{

        width:100%;

        border:none !important;

        border-radius:12px !important;

        background:rgba(255,255,255,0.08) !important;

        color:white !important;

        transition:0.2s;
    }

    .stButton button:hover,
    .stDownloadButton button:hover{

        transform:translateY(-2px);

        background:rgba(255,255,255,0.15) !important;
    }

    /* expander */

    .streamlit-expander{
        background:rgba(255,255,255,0.075) !important;
        border:1px solid rgba(255,255,255,0.13) !important;
        border-radius:16px !important;
        box-shadow:0 10px 24px rgba(0,0,0,0.16) !important;
        overflow:hidden !important;
    }

    .streamlit-expanderHeader{
        background:rgba(255,255,255,0.075) !important;
        border-radius:16px 16px 0 0 !important;
        min-height:42px !important;
        padding:0 14px !important;
        font-weight:800 !important;
    }

    .streamlit-expanderContent{
        background:rgba(255,255,255,0.025) !important;
        border-top:1px solid rgba(255,255,255,0.10) !important;
        padding:14px !important;
    }

    div[data-testid="stExpander"] {
        background:rgba(255,255,255,0.075) !important;
        border:1px solid rgba(255,255,255,0.13) !important;
        border-radius:16px !important;
        box-shadow:0 10px 24px rgba(0,0,0,0.16) !important;
        overflow:hidden !important;
    }

    div[data-testid="stExpander"] details {
        background:transparent !important;
        border:none !important;
    }

    div[data-testid="stExpander"] summary {
        background:rgba(255,255,255,0.075) !important;
        border-radius:16px 16px 0 0 !important;
        min-height:42px !important;
        padding:0 14px !important;
        font-weight:800 !important;
    }

    /* metric */

    [data-testid="metric-container"]{

        background:rgba(255,255,255,0.06);

        border:1px solid rgba(255,255,255,0.10);

        border-radius:20px;

        padding:22px;

        box-shadow:
        0 8px 24px rgba(0,0,0,0.15);

        backdrop-filter:blur(12px);
    }

    /* dataframe */

    [data-testid="stDataFrame"]{

        background:rgba(255,255,255,0.04) !important;

        border-radius:18px;

        overflow:hidden;

        border:1px solid rgba(255,255,255,0.08);

        backdrop-filter:blur(12px);
    }
            


    .glideDataEditor{

        background:rgba(255,255,255,0.03) !important;
    }

    .glideDataEditor *{

        background:transparent !important;

        color:white !important;
    }
    .kpi-card{

        background:rgba(255,255,255,0.05);

        border:1px solid rgba(255,255,255,0.08);

        border-radius:18px;

        padding:20px;

        text-align:center;

        min-height:70px;

        backdrop-filter:blur(10px);
    }

    .kpi-title{

        font-size:15px;

        color:#cfd8dc;

        margin-bottom:10px;
    }

    .kpi-value{

        font-size:24px;

        font-weight:700;

        color:white;
    }
            
    /* 구분선 */

    hr{
        border-color:rgba(255,255,255,0.08);
        margin:0.55rem 0 0.75rem 0 !important;
    }

    /* 다운로드 버튼 */

    .stDownloadButton button{

        width:100%;

        border:none !important;

        border-radius:12px !important;

        background:rgba(255,255,255,0.08) !important;

        color:white !important;

        transition:0.2s;
    }

    .stDownloadButton button:hover{

        transform:translateY(-2px);

        background:rgba(255,255,255,0.15) !important;
    }

    /* 데이터프레임 헤더 */

    .glideDataEditor [role="columnheader"]{

        background:rgba(255,255,255,0.08) !important;

        color:white !important;

        font-weight:700 !important;
    }
            
    /* ==========================
       AGGRID
       ========================== */

    .ag-root-wrapper {

        background: rgba(255,255,255,0.04) !important;

        border: 1px solid rgba(255,255,255,0.08) !important;

        border-radius: 18px !important;

        overflow: hidden !important;
    }

    .ag-header {

        background: rgba(255,255,255,0.08) !important;
    }

    .ag-header-cell-label {

        color: white !important;

        font-weight: 700 !important;
    }

    .ag-cell {

        background: transparent !important;

        color: white !important;
    }

    .ag-row {

        background: transparent !important;
    }

    .ag-row:hover {

        background: rgba(255,255,255,0.05) !important;
    }

    .ag-row-selected {

        background: rgba(59,130,246,0.30) !important;
    }

    .block-container{
        padding-top:0.55rem !important;
        padding-bottom:0.5rem !important;
        max-width:100% !important;
    }

    div[data-testid="stVerticalBlock"] {
        gap:0.46rem !important;
    }

    div[data-testid="stHorizontalBlock"] {
        gap:0.55rem !important;
        align-items:flex-start !important;
    }

    div[data-testid="column"] {
        align-self:flex-start !important;
    }

    .dashboard-box{

        background:rgba(255,255,255,0.05);

        border:1px solid rgba(255,255,255,0.10);

        border-radius:22px;

        padding:20px;

        backdrop-filter:blur(10px);

        height:100%;

        box-shadow:
        0 8px 20px rgba(0,0,0,0.15);
    }

    .dashboard-title{

        font-size:22px;

        font-weight:700;

        margin-bottom:15px;

        color:white;
    }

    .dashboard-page-title {
        display: flex;
        align-items: center;
        gap: 10px;
        font-size: 22px;
        font-weight: 800;
        line-height: 1;
        margin: 0;
        color: white;
    }

    .app-header-row {
        display: grid;
        grid-template-columns: 1.3fr 3fr;
        column-gap: 1rem;
        align-items: end;
        width: 100%;
        margin: 0 0 1.15rem 0;
    }

    .app-detail-page-title {
        margin: 0.05rem 0 0 0 !important;
    }

    .detail-header-actions {
        display: flex;
        align-items: center;
        gap: 0.85rem;
    }

    .app-subtitle {
        font-size:clamp(22px, 1.55vw, 28px);
        font-weight:700;
        line-height:1.2;
        margin:0.12rem 0 0 0;
        color:white;
    }

    .app-main-subtitle {
        margin: 0 0 clamp(0.65rem, 1vh, 1rem) 0 !important;
    }

    .search-result-title {
        font-size: 24px;
        font-weight: 800;
        line-height: 1.25;
        margin: 1.2rem 0 0.85rem 0;
        color: white;
    }

    .st-key-search_results_scroll {
        max-height: 360px !important;
        overflow-y: auto !important;
        overflow-x: hidden !important;
        padding-right: 4px !important;
    }

    .st-key-search_results_scroll div[data-testid="stVerticalBlock"] {
        gap: 0.44rem !important;
    }

    div[data-testid="stExpander"] [data-testid="stImage"] img {
        max-width: 100% !important;
        height: auto !important;
        object-fit: contain !important;
        margin-left: auto !important;
        margin-right: auto !important;
        display: block !important;
    }

    div[data-testid="stExpander"] [data-testid="stImage"] {
        display: flex !important;
        justify-content: center !important;
    }

    .dashboard-card-title {
        font-size: 18px;
        font-weight: 850;
        line-height: 1.25;
        margin: 0 0 8px 0;
        color: white;
        text-align: left;
    }

    .st-key-dashboard_card_recent .dashboard-card-title,
    .st-key-dashboard_card_top5 .dashboard-card-title {
        margin-bottom:7px;
    }

    .st-key-dashboard_card_top5 .dashboard-card-title {
        margin-bottom:8px;
    }

    div[data-testid="stMarkdownContainer"]:has(style) {
        display: none !important;
    }

    div[data-testid="stMarkdownContainer"]:has(.detail-title),
    div[data-testid="stMarkdownContainer"]:has(.detail-subtitle),
    div[data-testid="stMarkdownContainer"]:has(.detail-page-title),
    div[data-testid="stMarkdownContainer"]:has(.detail-field-label) {
        margin-bottom:0.15rem !important;
    }

    div[data-testid="stMarkdownContainer"] p {
        line-height:1.35 !important;
    }

    .st-key-dashboard_card_category,
    .st-key-dashboard_card_month,
    .st-key-dashboard_card_recent,
    .st-key-dashboard_card_top5 {
        width: 100% !important;
        background: rgba(255,255,255,0.075) !important;
        border: 1px solid rgba(255,255,255,0.13) !important;
        border-radius: 12px !important;
        box-shadow: 0 6px 16px rgba(0,0,0,0.14) !important;
        padding: 10px 12px !important;
        overflow: hidden !important;
    }

    .st-key-dashboard_card_category,
    .st-key-dashboard_card_month {
        min-height: var(--app-chart-card-h) !important;
        height: var(--app-chart-card-h) !important;
        max-height: var(--app-chart-card-h) !important;
        padding: 8px 10px 6px 10px !important;
    }

    .st-key-dashboard_card_recent,
    .st-key-dashboard_card_top5 {
        min-height: var(--app-list-card-h) !important;
        height: var(--app-list-card-h) !important;
        max-height: var(--app-list-card-h) !important;
        padding: 9px 10px !important;
    }

    div[class*="st-key-select_"] button {
        height: 42px !important;
        min-height: 42px !important;
        display: flex !important;
        padding: 0 12px !important;
        justify-content: flex-start !important;
        align-items: center !important;
        text-align: left !important;
        border-radius: 10px !important;
        box-shadow: none !important;
    }

    div[class*="st-key-select_"] button * {
        text-align: left !important;
        justify-content: flex-start !important;
    }

    div[class*="st-key-select_"] button div[data-testid="stMarkdownContainer"] {
        width: 100% !important;
        text-align: left !important;
    }

    div[class*="st-key-select_"] button p {
        width: 100% !important;
        text-align: left !important;
        white-space: nowrap !important;
        overflow: hidden !important;
        text-overflow: ellipsis !important;
        font-size: 14px !important;
        line-height: 1.2 !important;
    }

    div[class*="st-key-del_"] button {
        height: 42px !important;
        min-height: 42px !important;
        padding: 0 !important;
        border-radius: 10px !important;
        box-shadow: none !important;
    }

    div[class*="st-key-del_"] button p {
        text-align: center !important;
        font-size: 14px !important;
        line-height: 1 !important;
    }

    .st-key-save_case_btn,
    .st-key-excel_form_download_btn,
    .st-key-register_date_edit_btn,
    .st-key-edit_date_edit_btn,
    .st-key-back_dashboard_btn,
    .st-key-edit_back_btn,
    .st-key-detail_pdf_download_btn,
    .st-key-detail_edit_btn,
    .st-key-detail_delete_btn,
    .st-key-edit_save_btn,
    .st-key-edit_cancel_btn {
        width: fit-content !important;
    }

    .st-key-save_case_btn button,
    .st-key-excel_form_download_btn button,
    .st-key-register_date_edit_btn button,
    .st-key-edit_date_edit_btn button,
    .st-key-back_dashboard_btn button,
    .st-key-edit_back_btn button,
    .st-key-detail_pdf_download_btn button,
    .st-key-detail_edit_btn button,
    .st-key-detail_delete_btn button,
    .st-key-edit_save_btn button,
    .st-key-edit_cancel_btn button {
        width: auto !important;
        height: 32px !important;
        min-height: 32px !important;
        min-width: 86px !important;
        padding: 0 12px !important;
        border-radius: 8px !important;
        box-shadow: none !important;
        display: inline-flex !important;
        align-items: center !important;
        justify-content: center !important;
    }

    div[class*="st-key-save_case_btn"] div[data-testid="stButton"] button,
    div[class*="st-key-excel_form_download_btn"] div[data-testid="stDownloadButton"] button,
    div[class*="st-key-register_date_edit_btn"] div[data-testid="stButton"] button,
    div[class*="st-key-edit_date_edit_btn"] div[data-testid="stButton"] button,
    div[class*="st-key-back_dashboard_btn"] div[data-testid="stButton"] button,
    div[class*="st-key-edit_back_btn"] div[data-testid="stButton"] button,
    div[class*="st-key-detail_pdf_download_btn"] div[data-testid="stDownloadButton"] button,
    div[class*="st-key-detail_edit_btn"] div[data-testid="stButton"] button,
    div[class*="st-key-detail_delete_btn"] div[data-testid="stButton"] button,
    div[class*="st-key-edit_save_btn"] div[data-testid="stButton"] button,
    div[class*="st-key-edit_cancel_btn"] div[data-testid="stButton"] button,
    div[data-testid="stButton"][class*="st-key-save_case_btn"] button,
    div[data-testid="stDownloadButton"][class*="st-key-excel_form_download_btn"] button,
    div[data-testid="stButton"][class*="st-key-register_date_edit_btn"] button,
    div[data-testid="stButton"][class*="st-key-edit_date_edit_btn"] button,
    div[data-testid="stButton"][class*="st-key-back_dashboard_btn"] button,
    div[data-testid="stButton"][class*="st-key-edit_back_btn"] button,
    div[data-testid="stDownloadButton"][class*="st-key-detail_pdf_download_btn"] button,
    div[data-testid="stButton"][class*="st-key-detail_edit_btn"] button,
    div[data-testid="stButton"][class*="st-key-detail_delete_btn"] button,
    div[data-testid="stButton"][class*="st-key-edit_save_btn"] button,
    div[data-testid="stButton"][class*="st-key-edit_cancel_btn"] button {
        height: 32px !important;
        min-height: 32px !important;
        width: auto !important;
        min-width: 86px !important;
        padding: 0 12px !important;
        border-radius: 8px !important;
    }

    .st-key-save_case_btn button p,
    .st-key-excel_form_download_btn button p,
    .st-key-register_date_edit_btn button p,
    .st-key-edit_date_edit_btn button p,
    .st-key-back_dashboard_btn button p,
    .st-key-edit_back_btn button p,
    .st-key-detail_pdf_download_btn button p,
    .st-key-detail_edit_btn button p,
    .st-key-detail_delete_btn button p,
    .st-key-edit_save_btn button p,
    .st-key-edit_cancel_btn button p {
        font-size: 15px !important;
        line-height: 1 !important;
        text-align: center !important;
        white-space: nowrap !important;
        margin: 0 !important;
    }

    .st-key-back_dashboard_btn {
        margin: 0 0 0.04rem 0 !important;
    }

    .st-key-save_case_btn button,
    .st-key-excel_form_download_btn button {
        width: 112px !important;
        min-width: 112px !important;
        padding: 0 8px !important;
    }

    .st-key-detail_pdf_download_btn button,
    .st-key-detail_edit_btn button,
    .st-key-detail_delete_btn button {
        width: 126px !important;
        min-width: 126px !important;
        padding: 0 8px !important;
    }

    div[class*="_remove_saved_btn"],
    div[class*="_restore_btn"] {
        width: fit-content !important;
        margin: 0.15rem 0 0.25rem 0 !important;
    }

    div[class*="_remove_saved_btn"] button,
    div[class*="_restore_btn"] button {
        width: auto !important;
        min-width: 0 !important;
        height: 28px !important;
        min-height: 28px !important;
        padding: 0 10px !important;
        border-radius: 7px !important;
        display: inline-flex !important;
        align-items: center !important;
        justify-content: center !important;
        box-shadow: none !important;
    }

    div[class*="_remove_saved_btn"] button p,
    div[class*="_restore_btn"] button p {
        font-size: 13px !important;
        line-height: 1 !important;
        margin: 0 !important;
        white-space: nowrap !important;
    }

    .detail-title {
        font-size: 22px;
        font-weight: 800;
        line-height: 1.25;
        margin: 0.62rem 0 0.22rem 0;
    }

    .detail-subtitle {
        font-size: 20px;
        font-weight: 800;
        line-height: 1.25;
        margin: 0.45rem 0 0.28rem 0;
    }

    .detail-page-title {
        font-size: 22px;
        font-weight: 800;
        line-height: 1.25;
        margin: 0.18rem 0 0.28rem 0;
        white-space: nowrap;
    }

    .st-key-detail_header {
        margin: 0 0 0.08rem 0 !important;
        width: 100% !important;
    }

    .st-key-detail_header .detail-page-title {
        margin: 0 !important;
    }

    .st-key-detail_header [data-testid="stHorizontalBlock"] {
        align-items: center !important;
    }

    .st-key-detail_header [data-testid="column"] {
        display: flex !important;
        align-items: center !important;
    }

    .st-key-detail_info_card {
        width: 100% !important;
        background: rgba(255,255,255,0.06) !important;
        border: 1px solid rgba(255,255,255,0.12) !important;
        border-radius: 12px !important;
        padding: 14px 14px 24px 14px !important;
        margin: 0.1rem 0 0.08rem 0 !important;
        box-shadow: 0 6px 16px rgba(0,0,0,0.10) !important;
        overflow: hidden !important;
    }

    .st-key-detail_info_card [data-testid="stImage"] img {
        border-radius: 10px !important;
        max-height: 188px !important;
        object-fit: contain !important;
    }

    .detail-meta-grid {
        display:grid;
        grid-template-columns:repeat(2, minmax(0, 1fr));
        gap:12px;
        height:100%;
        align-content:start;
    }

    .detail-meta-cell {
        min-height:56px;
        border:1px solid rgba(255,255,255,0.10);
        border-radius:10px;
        padding:8px 10px;
        background:rgba(255,255,255,0.045);
    }

    .detail-meta-cell-wide {
        grid-column:1 / -1;
    }

    .detail-meta-label {
        display:block;
        font-size:11px;
        font-weight:800;
        line-height:1.2;
        margin-bottom:5px;
        color:rgba(255,255,255,0.66) !important;
    }

    .detail-meta-value {
        display:block;
        font-size:15px;
        font-weight:800;
        line-height:1.35;
        color:white !important;
        word-break:keep-all;
        overflow-wrap:anywhere;
    }

    .detail-section-grid {
        display:grid;
        grid-template-columns:repeat(2, minmax(0, 1fr));
        gap:10px;
        margin:0 0 0.95rem 0;
    }

    .detail-section-card {
        min-height:108px;
        border:1px solid rgba(255,255,255,0.12);
        border-left:4px solid #38bdf8;
        border-radius:12px;
        padding:12px 14px;
        background:rgba(255,255,255,0.055);
        box-shadow:0 6px 16px rgba(0,0,0,0.10);
    }

    .detail-section-card:nth-child(2) {
        border-left-color:#22c55e;
    }

    .detail-section-card:nth-child(3) {
        border-left-color:#f59e0b;
    }

    .detail-section-card:nth-child(4) {
        border-left-color:#f472b6;
    }

    .detail-section-title {
        display:flex;
        align-items:center;
        gap:8px;
        font-size:16px;
        font-weight:900;
        line-height:1.2;
        margin-bottom:7px;
    }

    .detail-section-title::before {
        content:"";
        width:7px;
        height:7px;
        border-radius:50%;
        background:currentColor;
        flex:0 0 auto;
    }

    .detail-section-body {
        font-size:14px;
        font-weight:600;
        line-height:1.38;
        color:rgba(255,255,255,0.88) !important;
        white-space:normal;
        word-break:keep-all;
        overflow-wrap:anywhere;
    }

    .detail-photo-title {
        font-size:15px;
        font-weight:900;
        line-height:1.18;
        margin:0.2rem 0 0.62rem 0;
    }

    .st-key-detail_case_photo,
    .st-key-detail_repair_photo {
        width: 100% !important;
        margin-bottom: 0.45rem !important;
    }

    .st-key-detail_case_photo > div,
    .st-key-detail_repair_photo > div {
        width: 100% !important;
    }

    .st-key-detail_case_photo [data-testid="stImage"],
    .st-key-detail_repair_photo [data-testid="stImage"] {
        width: 100% !important;
        height: 312px !important;
        border: 1px solid rgba(255,255,255,0.12) !important;
        border-radius: 10px !important;
        background: rgba(255,255,255,0.045) !important;
        padding: 6px !important;
        display: flex !important;
        align-items: center !important;
        justify-content: center !important;
    }

    .st-key-detail_case_photo [data-testid="stImage"] img,
    .st-key-detail_repair_photo [data-testid="stImage"] img {
        width: 100% !important;
        height: 300px !important;
        max-width: 100% !important;
        object-fit: contain !important;
        object-position: center center !important;
        border-radius: 8px !important;
    }

    .st-key-detail_case_photo [data-testid="stImage"] div,
    .st-key-detail_repair_photo [data-testid="stImage"] div {
        display: flex !important;
        justify-content: center !important;
    }

    @media (max-width: 900px) {
        .detail-meta-grid,
        .detail-section-grid {
            grid-template-columns:1fr;
        }
    }

    .detail-field-label {
        font-size: 19px;
        font-weight: 800;
        line-height: 1.2;
        margin: 0.24rem 0 0.08rem 0;
    }

    .detail-title + div[data-testid="stMarkdownContainer"] p,
    .detail-subtitle + div[data-testid="stMarkdownContainer"] p,
    .detail-field-label + div[data-testid="stMarkdownContainer"] p {
        line-height: 1.32 !important;
        margin-bottom: 0.28rem !important;
    }

    .detail-photo {
        margin: 0.05rem 0 0.3rem 0;
    }

    .st-key-top5_1 button,
    .st-key-top5_2 button,
    .st-key-top5_3 button,
    .st-key-top5_4 button,
    .st-key-top5_5 button {
        height: clamp(38px, calc(var(--app-dashboard-card-h) * 0.14), 46px) !important;
        min-height: clamp(38px, calc(var(--app-dashboard-card-h) * 0.14), 46px) !important;
        display: flex !important;
        justify-content: flex-start !important;
        align-items: center !important;
        text-align: left !important;
        padding: 0 10px !important;
        border: 1px solid rgba(255,255,255,0.10) !important;
        background: rgba(255,255,255,0.075) !important;
        box-shadow: none !important;
    }

    .st-key-top5_1 button *,
    .st-key-top5_2 button *,
    .st-key-top5_3 button *,
    .st-key-top5_4 button *,
    .st-key-top5_5 button * {
        text-align: left !important;
        justify-content: flex-start !important;
    }

    div[class*="st-key-top5_"] button {
        height: clamp(38px, calc(var(--app-dashboard-card-h) * 0.14), 46px) !important;
        min-height: clamp(38px, calc(var(--app-dashboard-card-h) * 0.14), 46px) !important;
        display: flex !important;
        justify-content: flex-start !important;
        align-items: center !important;
        text-align: left !important;
        padding: 0 10px !important;
        border-radius: 9px !important;
    }

    div[class*="st-key-top5_"] button * {
        text-align: left !important;
        justify-content: flex-start !important;
        min-height: 0 !important;
    }

    div[class*="st-key-top5_"] button div[data-testid="stMarkdownContainer"] {
        width: 100% !important;
        line-height: 1.2 !important;
    }

    div[class*="st-key-top5_"] button p {
        margin: 0 !important;
        line-height: 1.2 !important;
    }

    .st-key-top5_1 button div[data-testid="stMarkdownContainer"],
    .st-key-top5_2 button div[data-testid="stMarkdownContainer"],
    .st-key-top5_3 button div[data-testid="stMarkdownContainer"],
    .st-key-top5_4 button div[data-testid="stMarkdownContainer"],
    .st-key-top5_5 button div[data-testid="stMarkdownContainer"] {
        width: 100% !important;
        text-align: left !important;
    }

    .st-key-top5_1 button p,
    .st-key-top5_2 button p,
    .st-key-top5_3 button p,
    .st-key-top5_4 button p,
    .st-key-top5_5 button p {
        text-align: left !important;
        white-space: nowrap !important;
        overflow: hidden !important;
        text-overflow: ellipsis !important;
        font-size: clamp(13px, 0.74vw, 14px) !important;
        line-height: 1.2 !important;
    }

    .top5-meter {
        width:100%;
        margin: clamp(6px, calc(var(--app-dashboard-card-h) * 0.024), 9px) 0 clamp(9px, calc(var(--app-dashboard-card-h) * 0.032), 12px) 0;
    }

    .top5-track {
        width:100%;
        height: clamp(7px, calc(var(--app-dashboard-card-h) * 0.024), 9px);
        background:rgba(255,255,255,0.08);
        border-radius:10px;
        overflow:hidden;
    }

    .top5-fill {
        height:100%;
        border-radius:10px;
    }

    .top5-count {
        display:block;
        margin-top:2px;
        font-size:clamp(12px, 0.66vw, 13px);
        line-height:1.2;
        color:rgba(255,255,255,0.78) !important;
    }

    .st-key-dashboard_card_top5 div[data-testid="stVerticalBlock"] {
        gap:0.28rem !important;
    }

    .st-key-dashboard_card_top5 {
        display: flex !important;
        flex-direction: column !important;
    }

    .st-key-dashboard_card_top5 > div,
    .st-key-dashboard_card_top5 div[data-testid="stVerticalBlock"] {
        min-height: 0 !important;
    }

    .st-key-kpi_all button,
    .st-key-kpi_month button,
    .st-key-kpi_broken button,
    .st-key-kpi_defect button,
    .st-key-kpi_wrong button,
    .st-key-kpi_shortage button,
    .st-key-kpi_missing button,
    .st-key-kpi_etc button,
    .st-key-kpi_horizontal button,
    .st-key-kpi_welding button {
        width: 100% !important;
        height: var(--app-kpi-h) !important;
        min-height: var(--app-kpi-h) !important;
        padding: 8px 11px !important;
        border-radius: 12px !important;
        border: 1px solid rgba(255,255,255,0.13) !important;
        background: rgba(255,255,255,0.075) !important;
        box-shadow: 0 10px 24px rgba(0,0,0,0.16);
        display: flex !important;
        align-items: center !important;
        justify-content: center !important;
    }

    .st-key-kpi_all button div[data-testid="stMarkdownContainer"],
    .st-key-kpi_month button div[data-testid="stMarkdownContainer"],
    .st-key-kpi_broken button div[data-testid="stMarkdownContainer"],
    .st-key-kpi_defect button div[data-testid="stMarkdownContainer"],
    .st-key-kpi_wrong button div[data-testid="stMarkdownContainer"],
    .st-key-kpi_shortage button div[data-testid="stMarkdownContainer"],
    .st-key-kpi_missing button div[data-testid="stMarkdownContainer"],
    .st-key-kpi_etc button div[data-testid="stMarkdownContainer"],
    .st-key-kpi_horizontal button div[data-testid="stMarkdownContainer"],
    .st-key-kpi_welding button div[data-testid="stMarkdownContainer"] {
        width: 100% !important;
        display: flex !important;
        align-items: center !important;
        justify-content: center !important;
        text-align: center !important;
    }

    .st-key-kpi_all button p,
    .st-key-kpi_month button p,
    .st-key-kpi_broken button p,
    .st-key-kpi_defect button p,
    .st-key-kpi_wrong button p,
    .st-key-kpi_shortage button p,
    .st-key-kpi_missing button p,
    .st-key-kpi_etc button p,
    .st-key-kpi_horizontal button p,
    .st-key-kpi_welding button p {
        width: 100% !important;
        white-space: pre-line !important;
        text-align: center !important;
        font-size: 17px !important;
        font-weight: 700 !important;
        line-height: 1.28 !important;
        word-break: keep-all !important;
        overflow-wrap: normal !important;
        margin: 0 !important;
    }

    .st-key-kpi_all button:hover,
    .st-key-kpi_month button:hover,
    .st-key-kpi_broken button:hover,
    .st-key-kpi_defect button:hover,
    .st-key-kpi_wrong button:hover,
    .st-key-kpi_shortage button:hover,
    .st-key-kpi_missing button:hover,
    .st-key-kpi_etc button:hover,
    .st-key-kpi_horizontal button:hover,
    .st-key-kpi_welding button:hover {
        border-color: rgba(255,255,255,0.28) !important;
        background: rgba(255,255,255,0.14) !important;
    }

    /* Viewport fit: keep the current desktop feel while adapting to screen size. */
    :root {
        --app-max-width: none;
        --app-page-x: clamp(0.8rem, 1vw, 1.2rem);
        --app-page-y: clamp(2.55rem, 3.2vh, 3.1rem);
        --app-kpi-h: clamp(82px, 8.2vh, 96px);
        --app-chart-card-h: clamp(320px, 31vh, 350px);
        --app-list-card-h: clamp(310px, 30vh, 338px);
        --app-dashboard-card-h: var(--app-list-card-h);
        --app-chart-h: calc(var(--app-chart-card-h) - 36px);
        --app-grid-h: calc(var(--app-list-card-h) - 52px);
        --app-detail-photo-h: clamp(240px, 34vh, 312px);
        --app-search-h: clamp(236px, 28vh, 320px);
    }

    .block-container {
        width: 100% !important;
        max-width: none !important;
        margin-left: 0 !important;
        margin-right: 0 !important;
        padding-top: var(--app-page-y) !important;
        padding-left: var(--app-page-x) !important;
        padding-right: var(--app-page-x) !important;
        padding-bottom: 2rem !important;
    }

    [data-testid="stAppViewBlockContainer"],
    [data-testid="stMainBlockContainer"] {
        max-width:none !important;
        width:100% !important;
    }

    [data-testid="stMainBlockContainer"] {
        padding: var(--app-page-y) var(--app-page-x) 2rem var(--app-page-x) !important;
    }

    [data-testid="stAppViewBlockContainer"] {
        padding:0 !important;
        max-width:none !important;
        width:100% !important;
    }

    .st-key-search_results_scroll {
        min-height: var(--app-search-h) !important;
        max-height: var(--app-search-h) !important;
    }

    div[data-testid="stHorizontalBlock"]:has(.st-key-search_results_scroll) {
        align-items: stretch !important;
        gap: 0.75rem !important;
    }

    div[data-testid="stHorizontalBlock"]:has(.st-key-search_results_scroll) > div[data-testid="column"]:first-child {
        flex: 0 0 clamp(390px, 28%, 470px) !important;
        min-width: 390px !important;
        max-width: 29% !important;
    }

    div[data-testid="stHorizontalBlock"]:has(.st-key-search_results_scroll) > div[data-testid="column"]:last-child {
        flex: 1 1 71% !important;
        min-width: 0 !important;
    }

    div[data-testid="stHorizontalBlock"]:has(.st-key-dashboard_card_recent):has(.st-key-dashboard_card_top5) {
        align-items: stretch !important;
        gap: 0.75rem !important;
    }

    div[data-testid="stHorizontalBlock"]:has(.st-key-dashboard_card_recent):has(.st-key-dashboard_card_top5) > div[data-testid="column"]:first-child {
        flex: 0 0 calc(74% - 0.375rem) !important;
        max-width: calc(74% - 0.375rem) !important;
        min-width: 0 !important;
    }

    div[data-testid="stHorizontalBlock"]:has(.st-key-dashboard_card_recent):has(.st-key-dashboard_card_top5) > div[data-testid="column"]:last-child {
        flex: 0 0 calc(26% - 0.375rem) !important;
        max-width: calc(26% - 0.375rem) !important;
        min-width: 280px !important;
    }

    div[data-testid="stHorizontalBlock"]:has(.st-key-search_results_scroll) > div[data-testid="column"]:first-child input,
    div[data-testid="stHorizontalBlock"]:has(.st-key-search_results_scroll) > div[data-testid="column"]:first-child textarea,
    div[data-testid="stHorizontalBlock"]:has(.st-key-search_results_scroll) > div[data-testid="column"]:first-child div[data-baseweb="input"] {
        width: 100% !important;
    }

    .st-key-dashboard_card_category,
    .st-key-dashboard_card_month,
    .st-key-dashboard_card_recent,
    .st-key-dashboard_card_top5 {
        min-height: var(--app-list-card-h) !important;
        height: var(--app-list-card-h) !important;
        max-height: var(--app-list-card-h) !important;
    }

    .st-key-dashboard_card_category,
    .st-key-dashboard_card_month {
        min-height: var(--app-chart-card-h) !important;
        height: var(--app-chart-card-h) !important;
        max-height: var(--app-chart-card-h) !important;
    }

    .st-key-dashboard_card_category [data-testid="stPlotlyChart"],
    .st-key-dashboard_card_month [data-testid="stPlotlyChart"] {
        min-height: var(--app-chart-h) !important;
        height: var(--app-chart-h) !important;
        max-height: var(--app-chart-h) !important;
    }

    .st-key-dashboard_card_category [data-testid="stPlotlyChart"] > div,
    .st-key-dashboard_card_month [data-testid="stPlotlyChart"] > div {
        min-height: var(--app-chart-h) !important;
        height: var(--app-chart-h) !important;
        max-height: var(--app-chart-h) !important;
    }

    .st-key-dashboard_card_recent iframe {
        min-height: var(--app-grid-h) !important;
        height: var(--app-grid-h) !important;
        max-height: var(--app-grid-h) !important;
    }

    .st-key-detail_case_photo [data-testid="stImage"],
    .st-key-detail_repair_photo [data-testid="stImage"] {
        height: var(--app-detail-photo-h) !important;
    }

    .st-key-detail_case_photo [data-testid="stImage"] img,
    .st-key-detail_repair_photo [data-testid="stImage"] img {
        width: 100% !important;
        height: calc(var(--app-detail-photo-h) - 12px) !important;
        max-height: calc(var(--app-detail-photo-h) - 12px) !important;
        object-fit: contain !important;
        object-position: center center !important;
    }

    @media (min-width: 1600px) {
        :root {
            --app-max-width: none;
        }
    }

    @media (max-height: 850px) {
        :root {
            --app-page-y: 2.4rem;
            --app-kpi-h: 82px;
            --app-chart-card-h: 326px;
            --app-list-card-h: 306px;
            --app-dashboard-card-h: var(--app-list-card-h);
            --app-search-h: 240px;
        }
    }

    @media (max-height: 760px) {
        :root {
            --app-page-y: 2.1rem;
            --app-kpi-h: 78px;
            --app-chart-card-h: 310px;
            --app-list-card-h: 286px;
            --app-dashboard-card-h: var(--app-list-card-h);
            --app-detail-photo-h: 250px;
            --app-search-h: 220px;
        }

        .st-key-kpi_all button,
        .st-key-kpi_month button,
        .st-key-kpi_broken button,
        .st-key-kpi_defect button,
        .st-key-kpi_wrong button,
        .st-key-kpi_shortage button,
        .st-key-kpi_missing button,
        .st-key-kpi_etc button,
        .st-key-kpi_horizontal button,
        .st-key-kpi_welding button {
            height: 62px !important;
            min-height: 62px !important;
        }

        div[class*="st-key-top5_"] button {
            height: 38px !important;
            min-height: 38px !important;
        }
    }

    @media (max-width: 900px) {
        :root {
            --app-page-x: 0.85rem;
            --app-kpi-h: 92px;
            --app-chart-card-h: 330px;
            --app-list-card-h: 318px;
            --app-dashboard-card-h: var(--app-list-card-h);
            --app-detail-photo-h: 260px;
            --app-search-h: 240px;
        }

        .block-container {
            width: 100% !important;
            max-width: 100% !important;
        }

        div[data-testid="stHorizontalBlock"]:has(.st-key-dashboard_card_recent):has(.st-key-dashboard_card_top5) > div[data-testid="column"]:first-child,
        div[data-testid="stHorizontalBlock"]:has(.st-key-dashboard_card_recent):has(.st-key-dashboard_card_top5) > div[data-testid="column"]:last-child {
            flex: 1 1 100% !important;
            max-width: 100% !important;
            min-width: 0 !important;
        }
    }

    .st-key-main_excel_prepare_btn,
    div[class*="st-key-main_excel_prepare_btn"],
    .st-key-main_excel_download_btn,
    div[class*="st-key-main_excel_download_btn"] {
        position: static !important;
        margin: 0.35rem 0 0 var(--app-page-x) !important;
        width: auto !important;
        max-width: calc(100vw - (var(--app-page-x) * 2)) !important;
        z-index: auto !important;
    }

    .st-key-main_excel_prepare_btn div[data-testid="stButton"],
    div[class*="st-key-main_excel_prepare_btn"] div[data-testid="stButton"],
    .st-key-main_excel_download_btn div[data-testid="stDownloadButton"],
    div[class*="st-key-main_excel_download_btn"] div[data-testid="stDownloadButton"] {
        width: auto !important;
    }

    .st-key-main_excel_prepare_btn button,
    div[class*="st-key-main_excel_prepare_btn"] button,
    .st-key-main_excel_download_btn button,
    div[class*="st-key-main_excel_download_btn"] button {
        width: auto !important;
        min-width: 108px !important;
        min-height: 34px !important;
        height: 34px !important;
        padding: 0 14px !important;
        border-radius: 10px !important;
    }

    /* Responsive action buttons: prevent fixed-width buttons from overlapping in narrow layouts. */
    .st-key-save_case_btn,
    .st-key-excel_form_download_btn,
    .st-key-register_date_edit_btn,
    .st-key-edit_date_edit_btn,
    .st-key-back_dashboard_btn,
    .st-key-edit_back_btn,
    .st-key-detail_pdf_download_btn,
    .st-key-detail_edit_btn,
    .st-key-detail_delete_btn,
    .st-key-edit_save_btn,
    .st-key-edit_cancel_btn {
        width: 100% !important;
        min-width: 0 !important;
    }

    .st-key-save_case_btn div[data-testid="stButton"],
    .st-key-excel_form_download_btn div[data-testid="stDownloadButton"],
    .st-key-register_date_edit_btn div[data-testid="stButton"],
    .st-key-edit_date_edit_btn div[data-testid="stButton"],
    .st-key-back_dashboard_btn div[data-testid="stButton"],
    .st-key-edit_back_btn div[data-testid="stButton"],
    .st-key-detail_pdf_download_btn div[data-testid="stDownloadButton"],
    .st-key-detail_edit_btn div[data-testid="stButton"],
    .st-key-detail_delete_btn div[data-testid="stButton"],
    .st-key-edit_save_btn div[data-testid="stButton"],
    .st-key-edit_cancel_btn div[data-testid="stButton"] {
        width: 100% !important;
        min-width: 0 !important;
    }

    .st-key-save_case_btn button,
    .st-key-excel_form_download_btn button,
    .st-key-register_date_edit_btn button,
    .st-key-edit_date_edit_btn button,
    .st-key-back_dashboard_btn button,
    .st-key-edit_back_btn button,
    .st-key-detail_pdf_download_btn button,
    .st-key-detail_edit_btn button,
    .st-key-detail_delete_btn button,
    .st-key-edit_save_btn button,
    .st-key-edit_cancel_btn button {
        width: 100% !important;
        min-width: 0 !important;
        max-width: 100% !important;
        height: auto !important;
        min-height: 36px !important;
        padding: 7px 9px !important;
        display: flex !important;
        align-items: center !important;
        justify-content: center !important;
        box-sizing: border-box !important;
        overflow: hidden !important;
    }

    .st-key-save_case_btn button p,
    .st-key-excel_form_download_btn button p,
    .st-key-register_date_edit_btn button p,
    .st-key-edit_date_edit_btn button p,
    .st-key-back_dashboard_btn button p,
    .st-key-edit_back_btn button p,
    .st-key-detail_pdf_download_btn button p,
    .st-key-detail_edit_btn button p,
    .st-key-detail_delete_btn button p,
    .st-key-edit_save_btn button p,
    .st-key-edit_cancel_btn button p {
        width: 100% !important;
        margin: 0 !important;
        font-size: clamp(11px, 1.05vw, 15px) !important;
        line-height: 1.18 !important;
        text-align: center !important;
        white-space: normal !important;
        overflow-wrap: keep-all !important;
        word-break: keep-all !important;
    }

    .st-key-save_case_btn button p,
    .st-key-excel_form_download_btn button p,
    .st-key-register_date_edit_btn button p,
    .st-key-edit_date_edit_btn button p,
    .st-key-back_dashboard_btn button p,
    .st-key-edit_back_btn button p,
    .st-key-detail_pdf_download_btn button p,
    .st-key-detail_edit_btn button p,
    .st-key-detail_delete_btn button p,
    .st-key-edit_save_btn button p,
    .st-key-edit_cancel_btn button p {
        white-space: nowrap !important;
        overflow: visible !important;
        text-overflow: unset !important;
    }

    .st-key-save_case_btn button,
    .st-key-excel_form_download_btn button,
    .st-key-back_dashboard_btn button,
    .st-key-edit_back_btn button {
        min-height: 34px !important;
        padding-left: 10px !important;
        padding-right: 10px !important;
    }

    @media (max-width: 760px) {
        .st-key-save_case_btn button,
        .st-key-excel_form_download_btn button,
        .st-key-register_date_edit_btn button,
        .st-key-edit_date_edit_btn button,
        .st-key-back_dashboard_btn button,
        .st-key-edit_back_btn button,
        .st-key-detail_pdf_download_btn button,
        .st-key-detail_edit_btn button,
        .st-key-detail_delete_btn button,
        .st-key-edit_save_btn button,
        .st-key-edit_cancel_btn button {
            min-height: 40px !important;
            padding: 6px 7px !important;
        }
    }

    /* Action button rows */
    :root {
        --app-action-gap: clamp(6px, 0.45vw, 10px);
        --app-detail-action-gap: 48px;
        --app-detail-action-row-width: 496px;
        --app-detail-min-width: 560px;
        --app-action-btn-h: 36px;
    }

    html,
    body,
    [data-testid="stAppViewContainer"] {
        overflow-x: auto !important;
    }

    .block-container:has(.app-detail-page-title),
    .block-container:has(.st-key-detail_info_card),
    .block-container:has(div[class*="st-key-detail_info_card"]) {
        min-width: var(--app-detail-min-width) !important;
    }

    .st-key-register_action_row div[data-testid="stHorizontalBlock"],
    .st-key-detail_action_row div[data-testid="stHorizontalBlock"] {
        align-items: center !important;
        gap: var(--app-action-gap) !important;
    }

    .st-key-detail_action_row,
    div[class*="st-key-detail_action_row"] {
        width: max-content !important;
        min-width: var(--app-detail-action-row-width) !important;
        overflow: visible !important;
    }

    .st-key-detail_action_row div[data-testid="stHorizontalBlock"],
    div[class*="st-key-detail_action_row"] div[data-testid="stHorizontalBlock"] {
        display: flex !important;
        flex-wrap: nowrap !important;
        gap: 0 !important;
        width: max-content !important;
        min-width: var(--app-detail-action-row-width) !important;
        overflow: visible !important;
    }

    .st-key-detail_info_card,
    div[class*="st-key-detail_info_card"] {
        min-width: var(--app-detail-min-width) !important;
        overflow: visible !important;
    }

    .detail-action-gap {
        width: var(--app-detail-action-gap) !important;
        min-width: var(--app-detail-action-gap) !important;
        height: var(--app-action-btn-h) !important;
    }

    .st-key-register_action_row div[data-testid="column"],
    .st-key-detail_action_row div[data-testid="column"] {
        min-width: 0 !important;
    }

    .st-key-detail_action_row div[data-testid="stHorizontalBlock"] div[data-testid="column"],
    div[class*="st-key-detail_action_row"] div[data-testid="stHorizontalBlock"] div[data-testid="column"] {
        flex: 0 0 auto !important;
        margin: 0 !important;
        min-width: 0 !important;
        max-width: none !important;
    }

    .st-key-detail_action_row div[data-testid="stHorizontalBlock"] > div[data-testid="column"]:nth-child(1),
    div[class*="st-key-detail_action_row"] div[data-testid="stHorizontalBlock"] > div[data-testid="column"]:nth-child(1) {
        flex: 0 0 168px !important;
        width: 168px !important;
        min-width: 168px !important;
        max-width: 168px !important;
    }

    .st-key-detail_action_row div[data-testid="stHorizontalBlock"] > div[data-testid="column"]:nth-child(2),
    div[class*="st-key-detail_action_row"] div[data-testid="stHorizontalBlock"] > div[data-testid="column"]:nth-child(2),
    .st-key-detail_action_row div[data-testid="stHorizontalBlock"] > div[data-testid="column"]:nth-child(4),
    div[class*="st-key-detail_action_row"] div[data-testid="stHorizontalBlock"] > div[data-testid="column"]:nth-child(4) {
        flex: 0 0 var(--app-detail-action-gap) !important;
        width: var(--app-detail-action-gap) !important;
        min-width: var(--app-detail-action-gap) !important;
        max-width: var(--app-detail-action-gap) !important;
    }

    .st-key-detail_action_row div[data-testid="stHorizontalBlock"] > div[data-testid="column"]:nth-child(3),
    div[class*="st-key-detail_action_row"] div[data-testid="stHorizontalBlock"] > div[data-testid="column"]:nth-child(3),
    .st-key-detail_action_row div[data-testid="stHorizontalBlock"] > div[data-testid="column"]:nth-child(5),
    div[class*="st-key-detail_action_row"] div[data-testid="stHorizontalBlock"] > div[data-testid="column"]:nth-child(5) {
        flex: 0 0 116px !important;
        width: 116px !important;
        min-width: 116px !important;
        max-width: 116px !important;
    }

    .st-key-detail_action_row div[data-testid="stHorizontalBlock"] > div[data-testid="column"]:nth-child(6),
    div[class*="st-key-detail_action_row"] div[data-testid="stHorizontalBlock"] > div[data-testid="column"]:nth-child(6) {
        display: none !important;
        flex: 0 0 0 !important;
        width: 0 !important;
        min-width: 0 !important;
        max-width: 0 !important;
    }

    .st-key-register_action_row div[data-testid="stHorizontalBlock"] > div[data-testid="column"]:nth-child(1),
    .st-key-register_action_row div[data-testid="stHorizontalBlock"] > div[data-testid="column"]:nth-child(3) {
        flex: 0 0 auto !important;
        flex-basis: 102px !important;
        width: 102px !important;
        min-width: 102px !important;
    }

    .st-key-register_action_row div[data-testid="stHorizontalBlock"] > div[data-testid="column"]:nth-child(2) {
        flex: 0 0 var(--app-action-gap) !important;
        width: var(--app-action-gap) !important;
        min-width: var(--app-action-gap) !important;
    }

    .st-key-register_action_row button,
    .st-key-detail_action_row button {
        width: 100% !important;
        min-width: 0 !important;
        height: var(--app-action-btn-h) !important;
        min-height: var(--app-action-btn-h) !important;
        padding: 0 12px !important;
        box-sizing: border-box !important;
        white-space: nowrap !important;
    }

    .st-key-register_action_row button p,
    .st-key-detail_action_row button p {
        margin: 0 !important;
        font-size: 15px !important;
        line-height: 1 !important;
        white-space: nowrap !important;
        overflow: visible !important;
    }

    /* Fixed readable widths for action buttons. */
    .st-key-detail_pdf_download_btn,
    .st-key-detail_pdf_download_btn div[data-testid="stDownloadButton"],
    .st-key-detail_pdf_download_btn button {
        width: 168px !important;
        min-width: 168px !important;
        max-width: 168px !important;
    }

    .st-key-detail_edit_btn,
    .st-key-detail_delete_btn,
    .st-key-detail_edit_btn div[data-testid="stButton"],
    .st-key-detail_delete_btn div[data-testid="stButton"],
    .st-key-detail_edit_btn button,
    .st-key-detail_delete_btn button {
        width: 116px !important;
        min-width: 116px !important;
        max-width: 116px !important;
    }

    .st-key-save_case_btn,
    .st-key-excel_form_download_btn,
    .st-key-save_case_btn div[data-testid="stButton"],
    .st-key-excel_form_download_btn div[data-testid="stDownloadButton"],
    .st-key-save_case_btn button,
    .st-key-excel_form_download_btn button {
        width: 112px !important;
        min-width: 112px !important;
        max-width: 112px !important;
    }

    .st-key-detail_pdf_download_btn button p,
    .st-key-detail_edit_btn button p,
    .st-key-detail_delete_btn button p,
    .st-key-save_case_btn button p,
    .st-key-excel_form_download_btn button p {
        width: 100% !important;
        text-align: center !important;
        overflow: visible !important;
        text-overflow: unset !important;
    }

    .st-key-detail_pdf_download_btn,
    .st-key-detail_edit_btn,
    .st-key-save_case_btn {
        margin-right: clamp(6px, 0.45vw, 10px) !important;
    }

    .st-key-detail_action_row .st-key-detail_pdf_download_btn,
    .st-key-detail_action_row .st-key-detail_edit_btn,
    .st-key-detail_action_row .st-key-detail_delete_btn,
    div[class*="st-key-detail_action_row"] div[class*="st-key-detail_pdf_download_btn"],
    div[class*="st-key-detail_action_row"] div[class*="st-key-detail_edit_btn"],
    div[class*="st-key-detail_action_row"] div[class*="st-key-detail_delete_btn"] {
        margin-right: 0 !important;
    }

    @media (max-width: 900px) {
        .st-key-register_action_row div[data-testid="stHorizontalBlock"] {
            gap: 12px !important;
        }

        .st-key-detail_action_row div[data-testid="stHorizontalBlock"],
        div[class*="st-key-detail_action_row"] div[data-testid="stHorizontalBlock"] {
            gap: 0 !important;
        }

        .st-key-detail_pdf_download_btn,
        .st-key-detail_edit_btn,
        .st-key-save_case_btn {
            margin-right: 12px !important;
        }

        .st-key-detail_action_row .st-key-detail_pdf_download_btn,
        .st-key-detail_action_row .st-key-detail_edit_btn,
        .st-key-detail_action_row .st-key-detail_delete_btn,
        div[class*="st-key-detail_action_row"] div[class*="st-key-detail_pdf_download_btn"],
        div[class*="st-key-detail_action_row"] div[class*="st-key-detail_edit_btn"],
        div[class*="st-key-detail_action_row"] div[class*="st-key-detail_delete_btn"] {
            margin-right: 0 !important;
        }
    }

    </style>
    """, unsafe_allow_html=True)
    # ==========================
    # 사례번호 자동생성
    # ==========================

    if "register_case_date" not in st.session_state:
        st.session_state.register_case_date = datetime.now().date()

    if "register_date_edit_open" not in st.session_state:
        st.session_state.register_date_edit_open = False

    if "register_expander_open" not in st.session_state:
        st.session_state.register_expander_open = False

    case_id = get_next_case_id(st.session_state.register_case_date)

    if "selected_case" not in st.session_state:
        st.session_state.selected_case = None

    if "edit_case" not in st.session_state:
        st.session_state.edit_case = None

    if "dashboard_filter" not in st.session_state:
        st.session_state.dashboard_filter = None

    if "dashboard_month" not in st.session_state:
        st.session_state.dashboard_month = datetime.now().strftime("%Y%m")

    query_return_case_filter = st.query_params.get("return_case_filter")
    query_return_case_month = st.query_params.get("return_case_month")
    query_return_case_id = st.query_params.get("return_case_id")
    if isinstance(query_return_case_filter, list):
        query_return_case_filter = query_return_case_filter[0] if query_return_case_filter else ""
    if isinstance(query_return_case_month, list):
        query_return_case_month = query_return_case_month[0] if query_return_case_month else ""
    if isinstance(query_return_case_id, list):
        query_return_case_id = query_return_case_id[0] if query_return_case_id else ""

    if query_return_case_id:
        c.execute(
            """
            SELECT
                case_id,
                category,
                barcode,
                product,
                cause,
                action,
                repair_method,
                prevention,
                product_image,
                case_image,
                repair_image
            FROM cases
            WHERE case_id = ?
            """,
            (str(query_return_case_id),),
        )
        selected_query_case = c.fetchone()
        if selected_query_case:
            st.session_state.selected_case = selected_query_case
            st.session_state.edit_case = None
            st.session_state.dashboard_filter = None

    if query_return_case_filter and not query_return_case_id:
        st.session_state.dashboard_filter = query_return_case_filter
        st.session_state.selected_case = None
        st.session_state.edit_case = None

    if query_return_case_month and str(query_return_case_month).isdigit() and len(str(query_return_case_month)) == 6:
        st.session_state.dashboard_month = str(query_return_case_month)

    if st.session_state.dashboard_filter == "변심":
        st.session_state.dashboard_filter = "누락"

    if "register_form_version" not in st.session_state:
        st.session_state.register_form_version = 0

    page_title_html = ""
    is_detail_page = (
        st.session_state.selected_case is not None
        and st.session_state.edit_case is None
    )

    if not st.session_state.edit_case and not st.session_state.selected_case:
        page_title_html = ""

    main_column_ratio = [1.12, 2.88]
    main_header_grid = "1.12fr 2.88fr"

    if is_detail_page:
        header_left, header_right = st.columns(main_column_ratio, gap="small")

        with header_left:
            st.markdown(
                '<div class="app-subtitle">반품/AS 관리</div>',
                unsafe_allow_html=True,
            )

        with header_right:
            back_col, _back_spacer = st.columns([2.15, 7.85], gap="small")

            with back_col:
                if st.button(
                    "← 📊 대시보드",
                    key="back_dashboard_btn",
                    use_container_width=True
                ):
                    st.session_state.selected_case = None
                    st.rerun()

            st.markdown(
                '<div class="detail-page-title app-detail-page-title">사례 상세보기</div>',
                unsafe_allow_html=True,
            )

    elif st.session_state.edit_case:
        header_left, header_right = st.columns(main_column_ratio, gap="small")

        with header_left:
            st.markdown(
                '<div class="app-subtitle">반품/AS 관리</div>',
                unsafe_allow_html=True,
            )

        with header_right:
            back_col, _back_spacer = st.columns([2.15, 7.85], gap="small")

            with back_col:
                if st.button(
                    "← 📄 상세보기",
                    key="edit_back_btn",
                    use_container_width=True
                ):
                    edit_header_data = st.session_state.edit_case
                    clear_marker_session_state(f"edit_case_marker_{edit_header_data[0]}")
                    clear_marker_session_state(f"edit_repair_marker_{edit_header_data[0]}")

                    st.session_state.selected_case = edit_header_data
                    st.session_state.edit_case = None
                    st.session_state.edit_date_edit_open = False
                    st.session_state.edit_date_case_id = None

                    st.rerun()

            st.markdown(
                '<div class="detail-page-title app-detail-page-title">사례 수정</div>',
                unsafe_allow_html=True,
            )

    else:
        st.markdown(
            '<div class="app-subtitle app-main-subtitle">반품/AS 관리</div>',
            unsafe_allow_html=True,
        )

    # ==========================
    # 검색 화면
    # ==========================

    left, right = st.columns(main_column_ratio, gap="small")

    # --------------------------
    # 왼쪽
    # --------------------------
    with left:

        with st.expander(
            "사례 등록",
            expanded=st.session_state.register_expander_open
        ):

            register_form_version = st.session_state.register_form_version

            case_id_col, date_edit_col = st.columns([2.2, 1], gap="small")

            with case_id_col:

                st.text(f"사례번호 : {case_id}")

            with date_edit_col:

                if st.button(
                    "📅 일자 수정",
                    key="register_date_edit_btn"
                ):

                    st.session_state.register_expander_open = True
                    st.session_state.register_date_edit_open = (
                        not st.session_state.register_date_edit_open
                    )

            if st.session_state.register_date_edit_open:

                selected_register_date = st.date_input(
                    "발생일자",
                    value=st.session_state.register_case_date,
                    key=f"register_case_date_input_{register_form_version}"
                )

                if selected_register_date != st.session_state.register_case_date:

                    st.session_state.register_case_date = selected_register_date

                    st.rerun()

            category = st.selectbox(
                "유형",
                CASE_CATEGORIES,
                key=f"register_category_{register_form_version}"
            )

            product = st.text_input(
                "상품명",
                key=f"register_product_{register_form_version}"
            )

            barcode = st.text_input(
                "바코드",
                key=f"register_barcode_{register_form_version}"
            )

            product_image = st.file_uploader(
                "상품 대표사진",
                type=["jpg", "jpeg", "png"],
                key=f"register_product_image_{register_form_version}"
            )

            cause = st.text_area(
                "원인",
                height=120,
                key=f"register_cause_{register_form_version}"
            )

            case_image = st.file_uploader(
                "사례 첨부사진",
                type=["jpg", "jpeg", "png"],
                accept_multiple_files=True,
                key=f"register_case_image_{register_form_version}"
            )

            case_image_original_data = get_first_upload_bytes(case_image)

            case_image_data = part_marker_tool(
                case_image_original_data,
                f"register_case_marker_{register_form_version}",
                "사례사진 부위 표시",
                restore_image_data=case_image_original_data,
            )

            action = st.text_area(
                "조치방법",
                height=120,
                key=f"register_action_{register_form_version}"
            )

            repair_method = st.text_area(
                "수리방법",
                height=120,
                key=f"register_repair_method_{register_form_version}"
            )

            repair_image = st.file_uploader(
                "수리방법 사진",
                type=["jpg", "jpeg", "png"],
                accept_multiple_files=True,
                key=f"register_repair_image_{register_form_version}"
            )

            repair_image_original_data = get_first_upload_bytes(repair_image)

            repair_image_data = blue_circle_marker_tool(
                repair_image_original_data,
                f"register_repair_marker_{register_form_version}",
                show_photo_buttons=False,
            )

            prevention = st.text_area(
                "방지대책",
                height=120,
                key=f"register_prevention_{register_form_version}"
            )

            excel_form_upload = st.file_uploader(
                "엑셀 폼 업로드",
                type=["xlsx"],
                key=f"register_excel_form_{register_form_version}"
            )

            with st.container(key="register_action_row"):

                save_col, register_btn_gap, form_col = st.columns(
                    [1, 0.18, 1],
                    gap="small",
                )

                with save_col:

                    if st.button(
                        "💾 저장하기",
                        key="save_case_btn",
                        use_container_width=True
                    ):

                        if excel_form_upload is not None:
                            try:
                                registered_case_id = insert_case_from_excel_form(excel_form_upload)
                                st.session_state.register_form_version += 1
                                st.session_state.register_case_date = datetime.now().date()
                                st.session_state.register_date_edit_open = False
                                st.session_state.register_expander_open = False
                                st.success(f"엑셀 폼 저장 완료: {registered_case_id}")
                                st.rerun()
                            except Exception as exc:
                                st.error(f"엑셀 폼 저장 중 오류가 발생했습니다: {exc}")
                        else:
                            product_image_data = (
                                product_image.getvalue()
                                if product_image
                                else None
                            )

                            c.execute("""
                            INSERT INTO cases
                            (
                                case_id,
                                category,
                                barcode,
                                product,
                                cause,
                                action,
                                repair_method,
                                prevention,
                                product_image,
                                case_image,
                                case_image_original,
                                repair_image,
                                repair_image_original
                            )
                            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
                            """,
                            (
                                case_id,
                                category,
                                barcode,
                                product,
                                cause,
                                action,
                                repair_method,
                                prevention,
                                product_image_data,
                                case_image_data,
                                case_image_original_data,
                                repair_image_data,
                                repair_image_original_data
                            ))

                            conn.commit()

                            st.session_state.register_form_version += 1
                            st.session_state.register_case_date = datetime.now().date()
                            st.session_state.register_date_edit_open = False
                            st.session_state.register_expander_open = False

                            st.success("저장 완료")

                            st.rerun()

                with register_btn_gap:

                    st.markdown(
                        '<div class="action-button-gap"></div>',
                        unsafe_allow_html=True,
                    )

                with form_col:

                    st.download_button(
                        label="📗 폼다운",
                        data=create_case_excel_form(st.session_state.register_case_date),
                        file_name="사례등록_엑셀폼.xlsx",
                        mime=CASE_EXCEL_FORM_MIME,
                        key="excel_form_download_btn",
                        use_container_width=True
                    )

        if "top5_pending_keyword" in st.session_state:
            st.session_state["search_type"] = "상품명"
            st.session_state["search_keyword"] = st.session_state.pop("top5_pending_keyword")
            st.session_state["top5_keyword"] = st.session_state["search_keyword"]
            st.session_state.dashboard_filter = None
            st.session_state.selected_case = None

        search_type = st.radio(
            "검색 기준",
            ["상품명", "바코드", "사례번호", "유형", "원인", "수리방법"],
            horizontal=True,
            key="search_type"
        )

        keyword = st.text_input(
            "검색어",
            key="search_keyword"
        )

        st.markdown(
            '<div class="search-result-title">검색 결과</div>',
            unsafe_allow_html=True
        )

    results = []

    dashboard_filter = st.session_state.get(
        "dashboard_filter",
        None
    )

    if dashboard_filter == "ALL":

        c.execute("""
        SELECT case_id, category, product
        FROM cases
        ORDER BY id DESC
        """)

        results = c.fetchall()

    elif dashboard_filter == "MONTH":

        current_month = st.session_state.get("dashboard_month") or datetime.now().strftime("%Y%m")

        c.execute("""
        SELECT case_id, category, product
        FROM cases
        WHERE substr(case_id,1,6)=?
        ORDER BY id DESC
        """, (current_month,))

        results = c.fetchall()

    elif dashboard_filter:

        c.execute("""
        SELECT case_id, category, product
        FROM cases
        WHERE category=?
        ORDER BY id DESC
        """, (dashboard_filter,))

        results = c.fetchall()

    elif keyword:

        if search_type == "상품명":

            c.execute("""
            SELECT case_id, category, product
            FROM cases
            WHERE product LIKE ?
            """, (f"%{keyword}%",))

        elif search_type == "바코드":

            c.execute("""
            SELECT case_id, category, product
            FROM cases
            WHERE barcode LIKE ?
            """, (f"%{keyword}%",))

        elif search_type == "사례번호":

            c.execute("""
            SELECT case_id, category, product
            FROM cases
            WHERE case_id LIKE ?
            """, (f"%{keyword}%",))

        elif search_type == "유형":

            c.execute("""
            SELECT case_id, category, product
            FROM cases
            WHERE category LIKE ?
            """, (f"%{keyword}%",))

        elif search_type == "원인":

            c.execute("""
            SELECT case_id, category, product
            FROM cases
            WHERE cause LIKE ?
            """, (f"%{keyword}%",))

        elif search_type == "수리방법":

            c.execute("""
            SELECT case_id, category, product
            FROM cases
            WHERE repair_method LIKE ?
            """, (f"%{keyword}%",))

        results = c.fetchall()

    if results:

        with left.container(
            height=320,
            border=False,
            key="search_results_scroll"
        ):

            for row in results:

                product_name = row[2]

                if len(product_name) > 45:
                    product_name = product_name[:45] + "..."

                row_col1, row_col2 = st.columns([18, 3], gap="small")

                with row_col1:

                    if st.button(
                        f"{row[0]} | {row[1]} | {product_name}",
                        key=f"select_{row[0]}",
                        use_container_width=True
                    ):

                        c.execute("""
                        SELECT
                            case_id,
                            category,
                            barcode,
                            product,
                            cause,
                            action,
                            repair_method,
                            prevention,
                            product_image,
                            case_image,
                            repair_image
                        FROM cases
                        WHERE case_id = ?
                        """, (row[0],))

                        st.session_state.selected_case = c.fetchone()

                        st.rerun()

                with row_col2:

                    if st.button(
                        "❌",
                        key=f"del_{row[0]}",
                        use_container_width=True
                    ):

                        c.execute(
                            "DELETE FROM cases WHERE case_id = ?",
                            (row[0],)
                        )

                        conn.commit()

                        st.success("삭제 완료")

                        st.rerun()

    def show_dashboard():

        # =====================
        # 통계
        # =====================

        c.execute("SELECT COUNT(*) FROM cases")
        total_count = c.fetchone()[0]

        c.execute("SELECT COUNT(*) FROM cases WHERE category='파손'")
        broken_count = c.fetchone()[0]

        c.execute("SELECT COUNT(*) FROM cases WHERE category='불량'")
        defect_count = c.fetchone()[0]

        c.execute("SELECT COUNT(*) FROM cases WHERE category='오발송'")
        wrong_count = c.fetchone()[0]

        c.execute("SELECT COUNT(*) FROM cases WHERE category='쇼트'")
        shortage_count = c.fetchone()[0]

        this_month = st.session_state.get("dashboard_month") or datetime.now().strftime("%Y%m")

        c.execute("""
        SELECT COUNT(*)
        FROM cases
        WHERE substr(case_id,1,6)=?
        """, (this_month,))
        month_count = c.fetchone()[0]

            # =====================
        # KPI 통계
        # =====================

        c.execute("SELECT COUNT(*) FROM cases WHERE category='누락'")
        missing_count = c.fetchone()[0]

        c.execute("SELECT COUNT(*) FROM cases WHERE category='기타'")
        etc_count = c.fetchone()[0]

        c.execute("SELECT COUNT(*) FROM cases WHERE category='수평'")
        horizontal_count = c.fetchone()[0]

        c.execute("SELECT COUNT(*) FROM cases WHERE category='용접'")
        welding_count = c.fetchone()[0]

        active_kpi_key = {
            "ALL": "kpi_all",
            "MONTH": "kpi_month",
            "파손": "kpi_broken",
            "불량": "kpi_defect",
            "오발송": "kpi_wrong",
            "쇼트": "kpi_shortage",
            "누락": "kpi_missing",
            "기타": "kpi_etc",
            "수평": "kpi_horizontal",
            "용접": "kpi_welding",
        }.get(st.session_state.get("dashboard_filter"))

        # =====================
        # KPI 버튼
        # =====================

        k1, k2, k3, k4, k5 = st.columns(5, gap="small")

        with k1:

            if st.button(
                f"📄 전체 사례\u00a0\u00a0{total_count:,}건",
                key="kpi_all",
                use_container_width=True,
                type="primary" if active_kpi_key == "kpi_all" else "secondary"
            ):

                st.session_state.selected_case = None
                st.session_state.dashboard_filter = "ALL"

                st.rerun()

        with k2:

            if st.button(
                f"📅 이번 달\u00a0\u00a0{month_count:,}건",
                key="kpi_month",
                use_container_width=True,
                type="primary" if active_kpi_key == "kpi_month" else "secondary"
            ):

                st.session_state.selected_case = None
                st.session_state.dashboard_filter = "MONTH"
                st.session_state.dashboard_month = datetime.now().strftime("%Y%m")

                st.rerun()

        with k3:

            if st.button(
                f"⚠️ 파손\u00a0\u00a0{broken_count:,}건",
                key="kpi_broken",
                use_container_width=True,
                type="primary" if active_kpi_key == "kpi_broken" else "secondary"
            ):

                st.session_state.selected_case = None
                st.session_state.dashboard_filter = "파손"

                st.rerun()

        with k4:

            if st.button(
                f"🛠️ 불량\u00a0\u00a0{defect_count:,}건",
                key="kpi_defect",
                use_container_width=True,
                type="primary" if active_kpi_key == "kpi_defect" else "secondary"
            ):

                st.session_state.selected_case = None
                st.session_state.dashboard_filter = "불량"

                st.rerun()

        with k5:

            if st.button(
                f"🚚 오발송\u00a0\u00a0{wrong_count:,}건",
                key="kpi_wrong",
                use_container_width=True,
                type="primary" if active_kpi_key == "kpi_wrong" else "secondary"
            ):

                st.session_state.selected_case = None
                st.session_state.dashboard_filter = "오발송"

                st.rerun()

        k6, k7, k8, k9, k10 = st.columns(5, gap="small")

        with k6:

            if st.button(
                f"📦 쇼트\u00a0\u00a0{shortage_count:,}건",
                key="kpi_shortage",
                use_container_width=True,
                type="primary" if active_kpi_key == "kpi_shortage" else "secondary"
            ):

                st.session_state.selected_case = None
                st.session_state.dashboard_filter = "쇼트"

                st.rerun()

        with k7:

            if st.button(
                f"🧩 누락\u00a0\u00a0{missing_count:,}건",
                key="kpi_missing",
                use_container_width=True,
                type="primary" if active_kpi_key == "kpi_missing" else "secondary"
            ):

                st.session_state.selected_case = None
                st.session_state.dashboard_filter = "누락"

                st.rerun()

        with k8:

            if st.button(
                f"📁 기타\u00a0\u00a0{etc_count:,}건",
                key="kpi_etc",
                use_container_width=True,
                type="primary" if active_kpi_key == "kpi_etc" else "secondary"
            ):

                st.session_state.selected_case = None
                st.session_state.dashboard_filter = "기타"

                st.rerun()

        with k9:

            if st.button(
                f"📏 수평\u00a0\u00a0{horizontal_count:,}건",
                key="kpi_horizontal",
                use_container_width=True,
                type="primary" if active_kpi_key == "kpi_horizontal" else "secondary"
            ):

                st.session_state.selected_case = None
                st.session_state.dashboard_filter = "수평"

                st.rerun()

        with k10:

            if st.button(
                f"🔥 용접\u00a0\u00a0{welding_count:,}건",
                key="kpi_welding",
                use_container_width=True,
                type="primary" if active_kpi_key == "kpi_welding" else "secondary"
            ):

                st.session_state.selected_case = None
                st.session_state.dashboard_filter = "용접"

                st.rerun()

        # =====================
        # 차트
        # =====================

        chart1, chart2 = st.columns([1, 1], gap="small")

        # ---------------------
        # 유형별 발생 현황
        # ---------------------

        with chart1:

            with st.container(key="dashboard_card_category"):

                st.markdown(
                    '<div class="dashboard-card-title">📉 유형별 발생 현황</div>',
                    unsafe_allow_html=True
                )

                c.execute("""
                SELECT category, COUNT(*)
                FROM cases
                GROUP BY category
                """)

                pie_df = pd.DataFrame(
                    c.fetchall(),
                    columns=["유형", "건수"]
                )

                if not pie_df.empty:

                    import plotly.express as px

                    fig = px.pie(
                        pie_df,
                        names="유형",
                        values="건수",
                        hole=0.60,
                        color_discrete_sequence=[
                            "#ef4444",
                            "#f97316",
                            "#eab308",
                            "#22c55e",
                            "#3b82f6"
                        ]
                    )

                    fig.update_layout(
                        height=315,
                        margin=dict(
                            l=10,
                            r=10,
                            t=4,
                            b=6
                        ),
                        paper_bgcolor="rgba(0,0,0,0)",
                        plot_bgcolor="rgba(0,0,0,0)",
                        font_color="white",
                        legend=dict(
                            font=dict(size=12),
                            orientation="v",
                            x=0.84,
                            y=0.5,
                            xanchor="left",
                            yanchor="middle"
                        )
                    )

                    fig.update_traces(
                        textinfo="percent",
                        textfont_size=12,
                        domain=dict(
                            x=[0.03, 0.78],
                            y=[0.08, 0.92]
                        )
                    )

                    st.plotly_chart(
                        fig,
                        use_container_width=True
                    )

        # ---------------------
        # 발생 추이
        # ---------------------

        with chart2:

            with st.container(key="dashboard_card_month"):

                trend_year = str(date.today().year)

                st.markdown(
                    f'<div class="dashboard-card-title">📈 {trend_year}년 발생추이</div>',
                    unsafe_allow_html=True
                )

                c.execute("""
                SELECT
                    substr(case_id,1,6) AS month,
                    COUNT(*) AS cnt
                FROM cases
                WHERE substr(case_id,1,4) = ?
                GROUP BY month
                ORDER BY month
                """, (trend_year,))

                trend_counts = dict(c.fetchall())
                trend_df = pd.DataFrame(
                    {
                        "월": [f"{month}월" for month in range(1, 13)],
                        "건수": [
                            trend_counts.get(f"{trend_year}{month:02d}", 0)
                            for month in range(1, 13)
                        ],
                    }
                )

                if not trend_df.empty:

                    import plotly.graph_objects as go

                    fig = go.Figure()

                    dates = trend_df["월"].tolist()
                    counts = trend_df["건수"].tolist()
                    max_count = max(counts)
                    y_dtick = max(1, math.ceil(max_count / 4))
                    y_axis_top = max(
                        y_dtick,
                        math.ceil((max_count + y_dtick) / y_dtick) * y_dtick
                    )
                    count_labels = [count if count > 0 else "" for count in counts]

                    if len(trend_df) == 1:
                        fig.add_trace(
                            go.Scatter(
                                x=dates,
                                y=counts,
                                mode="markers+text",
                                text=count_labels,
                                textposition="top center",
                                textfont=dict(size=13),
                                marker=dict(
                                    color="#60a5fa",
                                    size=9
                                ),
                                cliponaxis=False,
                                hovertemplate="%{x}<br>%{y}건<extra></extra>"
                            )
                        )
                    else:
                        for i in range(1, len(trend_df)):
                            before = counts[i - 1]
                            after = counts[i]

                            if after > before:
                                line_color = "#ef4444"
                            elif after < before:
                                line_color = "#60a5fa"
                            else:
                                line_color = "#94a3b8"

                            fig.add_trace(
                                go.Scatter(
                                    x=[dates[i - 1], dates[i]],
                                    y=[before, after],
                                    mode="lines",
                                    line=dict(
                                        color=line_color,
                                        width=3
                                    ),
                                    hovertemplate="%{x}<br>%{y}건<extra></extra>"
                                )
                            )

                        fig.add_trace(
                            go.Scatter(
                                x=dates,
                                y=counts,
                                mode="markers+text",
                                text=count_labels,
                                textposition="top center",
                                textfont=dict(size=13),
                                marker=dict(
                                    color="white",
                                    size=9,
                                    line=dict(
                                        color="#0f766e",
                                        width=2
                                    )
                                ),
                                cliponaxis=False,
                                hovertemplate="%{x}<br>%{y}건<extra></extra>"
                            )
                        )

                    fig.update_layout(
                        height=315,
                        margin=dict(
                            l=12,
                            r=10,
                            t=20,
                            b=8
                        ),
                        paper_bgcolor="rgba(0,0,0,0)",
                        plot_bgcolor="rgba(0,0,0,0)",
                        font_color="white",
                        font=dict(size=13),
                        hovermode="x unified",
                        showlegend=False,
                        xaxis_title="",
                        yaxis_title="",
                        xaxis=dict(
                            type="category",
                            tickangle=0,
                            tickfont=dict(size=12)
                        ),
                        yaxis=dict(
                            tickmode="linear",
                            dtick=y_dtick,
                            tickformat="d",
                            tickfont=dict(size=12),
                            range=[
                                0,
                                y_axis_top
                            ]
                        )
                    )

                    st.plotly_chart(
                        fig,
                        use_container_width=True
                    )

        # =====================
        # 하단영역
        # =====================

        lower_left, lower_right = st.columns([3, 1], gap="small")

        # ---------------------
        # 최근 등록 사례
        # ---------------------

        with lower_left:

            with st.container(key="dashboard_card_recent"):

                st.markdown(
                    '<div class="dashboard-card-title">📜 최근 등록 사례</div>',
                    unsafe_allow_html=True
                )

                c.execute("""
                SELECT
                    case_id,
                    category,
                    product
                FROM cases
                ORDER BY case_id DESC
                """)

                recent_df = pd.DataFrame(
                    c.fetchall(),
                    columns=["사례번호", "유형", "상품명"]
                )

                gb = GridOptionsBuilder.from_dataframe(recent_df)

                gb.configure_default_column(
                    cellStyle={
                        "color": "white",
                        "backgroundColor": "#073b36",
                        "textAlign": "center"
                    }
                )

                gb.configure_grid_options(
                    rowStyle={
                        "backgroundColor": "#073b36",
                        "color": "white"
                    },
                    rowHeight=28,
                    headerHeight=34,
                    suppressHorizontalScroll=True
                )
                gb.configure_column(
                    "사례번호",
                    width=110,
                    minWidth=110
                )

                gb.configure_column(
                    "유형",
                    width=80
                )

                gb.configure_column(
                    "상품명",
                    flex=1
                )
                gb.configure_selection(
                    selection_mode="single",
                    use_checkbox=False
                )

                grid_options = gb.build()

                grid_response = AgGrid(
                    recent_df,
                    gridOptions=grid_options,
                    height=276,
                    theme="streamlit",
                    fit_columns_on_grid_load=True,

                    custom_css={
                        ".ag-root-wrapper": {
                            "background-color": "#052c28 !important"
                        },

                        ".ag-header": {
                            "background-color": "rgba(255,255,255,0.10) !important"
                        },

                        ".ag-header-cell-label": {
                            "color": "white !important",
                            "font-weight": "700 !important",
                            "justify-content": "center !important"
                        },

                        ".ag-cell": {
                            "background-color": "#073b36 !important",
                            "color": "white !important",
                            "font-size": "13px !important",
                            "line-height": "28px !important"
                        },
                        ".ag-row-even .ag-cell": {
                            "background-color": "#073b36 !important"
                        },
                        ".ag-row-odd .ag-cell": {
                            "background-color": "#06433d !important"
                        },
                        ".ag-center-cols-container": {
                            "background-color": "#052c28 !important"
                        },

                        ".ag-center-cols-viewport": {
                            "background-color": "#052c28 !important"
                        },

                        ".ag-body-viewport": {
                            "background-color": "#052c28 !important"
                        },

                        ".ag-body-horizontal-scroll": {
                            "display": "none !important",
                            "height": "0 !important",
                            "min-height": "0 !important"
                        },

                        ".ag-body-horizontal-scroll-viewport": {
                            "display": "none !important",
                            "height": "0 !important",
                            "min-height": "0 !important"
                        },

                        ".ag-row": {
                            "background-color": "#073b36 !important",
                            "color": "white !important"
                        },

                        ".ag-row-hover": {
                            "background-color": "#0d6b63 !important"
                        },
                        ".ag-row-hover .ag-cell": {
                            "background-color": "#0d6b63 !important"
                        }
                    }
                )

                selected_rows = grid_response.get("selected_rows", [])

                if selected_rows is not None and len(selected_rows) > 0:

                    if isinstance(selected_rows, pd.DataFrame):

                        case_id = selected_rows.iloc[0]["사례번호"]

                    else:

                        case_id = selected_rows[0]["사례번호"]

                    c.execute("""
                    SELECT
                        case_id,
                        category,
                        barcode,
                        product,
                        cause,
                        action,
                        repair_method,
                        prevention,
                        product_image,
                        case_image,
                        repair_image
                    FROM cases
                    WHERE case_id = ?
                    """, (case_id,))

                    st.session_state.selected_case = c.fetchone()

                    st.rerun()
        # ---------------------
        # TOP5
        # ---------------------

        with lower_right:

            with st.container(key="dashboard_card_top5"):

                st.markdown(
                    '<div class="dashboard-card-title">💥 자주 발생하는 상품 TOP3</div>',
                    unsafe_allow_html=True
                )

                c.execute("""
                SELECT
                    product,
                    COUNT(*) cnt
                FROM cases
                GROUP BY product
                ORDER BY cnt DESC
                LIMIT 3
                """)

                top_products = c.fetchall()

                if top_products:

                    max_count = top_products[0][1]

                    colors = [
                        "#ef4444",
                        "#f97316",
                        "#eab308",
                        "#22c55e",
                        "#3b82f6"
                    ]

                    for idx, row in enumerate(top_products, start=1):

                        product = row[0]
                        count = row[1]

                        percent = (count / max_count) * 100

                        if st.button(
                            f"{idx}. {product}",
                            key=f"top5_{idx}",
                            use_container_width=True
                        ):

                            st.session_state["top5_pending_keyword"] = product
                            st.session_state.dashboard_filter = None
                            st.session_state.selected_case = None

                            st.rerun()

                        st.markdown(
                            f"""
                            <div class="top5-meter">
                                <div class="top5-track">
                                    <div
                                        class="top5-fill"
                                        style="width:{percent}%; background:{colors[idx-1]};"
                                    ></div>
                                </div>
                                <span class="top5-count">{count}건</span>
                            </div>
                            """,
                            unsafe_allow_html=True
                        )
    # --------------------------
    # 상세보기
    # --------------------------

    def show_detail():

        data = st.session_state.selected_case

        with st.container(key="detail_info_card"):

            col1, col2 = st.columns([0.85, 3.15], gap="small")

            with col1:

                if data[8]:

                    img = Image.open(BytesIO(data[8]))

                    img = ImageOps.exif_transpose(img)

                    st.image(
                        fit_detail_photo(img),
                        use_container_width=True
                    )

            with col2:

                st.markdown(
                    f"""
                    <div class="detail-meta-grid">
                        <div class="detail-meta-cell">
                            <span class="detail-meta-label">사례번호</span>
                            <span class="detail-meta-value">{detail_html(data[0])}</span>
                        </div>
                        <div class="detail-meta-cell">
                            <span class="detail-meta-label">유형</span>
                            <span class="detail-meta-value">{detail_html(data[1])}</span>
                        </div>
                        <div class="detail-meta-cell detail-meta-cell-wide">
                            <span class="detail-meta-label">상품명</span>
                            <span class="detail-meta-value">{detail_html(data[3])}</span>
                        </div>
                        <div class="detail-meta-cell detail-meta-cell-wide">
                            <span class="detail-meta-label">바코드</span>
                            <span class="detail-meta-value">{detail_html(data[2])}</span>
                        </div>
                    </div>
                    """,
                    unsafe_allow_html=True
                )

        st.markdown(
            f"""
            <div class="detail-section-grid">
                <div class="detail-section-card">
                    <div class="detail-section-title">원인</div>
                    <div class="detail-section-body">{detail_html(data[4])}</div>
                </div>
                <div class="detail-section-card">
                    <div class="detail-section-title">조치방법</div>
                    <div class="detail-section-body">{detail_html(data[5])}</div>
                </div>
                <div class="detail-section-card">
                    <div class="detail-section-title">수리방법</div>
                    <div class="detail-section-body">{detail_html(data[6])}</div>
                </div>
                <div class="detail-section-card">
                    <div class="detail-section-title">방지대책</div>
                    <div class="detail-section-body">{detail_html(data[7])}</div>
                </div>
            </div>
            """,
            unsafe_allow_html=True
        )

        if data[9] and data[10]:

            photo_col1, photo_col2 = st.columns(2, gap="small")

            with photo_col1:

                if data[9]:

                    with st.container(key="detail_case_photo"):

                        st.markdown(
                            '<div class="detail-photo-title">사례 사진</div>',
                            unsafe_allow_html=True
                        )

                        img = fit_detail_photo(Image.open(BytesIO(data[9])))

                        st.image(
                            img,
                            use_container_width=True
                        )

            with photo_col2:

                if data[10]:

                    with st.container(key="detail_repair_photo"):

                        st.markdown(
                            '<div class="detail-photo-title">수리 사진</div>',
                            unsafe_allow_html=True
                        )

                        img = fit_detail_photo(Image.open(BytesIO(data[10])))

                        st.image(
                            img,
                            use_container_width=True
                        )

        elif data[9]:

            photo_col1, _photo_spacer = st.columns(2, gap="small")

            with photo_col1:

                with st.container(key="detail_case_photo"):

                    st.markdown(
                        '<div class="detail-photo-title">사례 사진</div>',
                        unsafe_allow_html=True
                    )

                    img = fit_detail_photo(Image.open(BytesIO(data[9])))

                    st.image(
                        img,
                        use_container_width=True
                    )

        elif data[10]:

            _photo_spacer, photo_col2 = st.columns(2, gap="small")

            with photo_col2:

                with st.container(key="detail_repair_photo"):

                    st.markdown(
                        '<div class="detail-photo-title">수리 사진</div>',
                        unsafe_allow_html=True
                    )

                    img = fit_detail_photo(Image.open(BytesIO(data[10])))

                    st.image(
                        img,
                        use_container_width=True
                    )

        with st.container(key="detail_action_row"):

            pdf_btn, _detail_gap1, btn1, _detail_gap2, btn2, _detail_button_spacer = st.columns(
                [168, 48, 116, 48, 116, 1],
                gap="small",
            )

            with pdf_btn:

                st.download_button(
                    label="📄 PDF 다운로드",
                    data=create_detail_pdf(data),
                    file_name=f"{data[0]}_사례상세보기.pdf",
                    mime="application/pdf",
                    key="detail_pdf_download_btn",
                    use_container_width=True
                )

            with _detail_gap1:

                st.markdown(
                    '<div class="detail-action-gap"></div>',
                    unsafe_allow_html=True,
                )

            with btn1:

                if st.button(
                    "✏️ 수정하기",
                    key="detail_edit_btn",
                    use_container_width=True
                ):

                    clear_marker_session_state(
                        f"edit_case_marker_{data[0]}",
                        include_canvas_state=True,
                    )
                    clear_marker_session_state(
                        f"edit_repair_marker_{data[0]}",
                        include_canvas_state=True,
                    )

                    st.session_state.marker_canvas_session_version = (
                        st.session_state.get("marker_canvas_session_version", 0) + 1
                    )

                    st.session_state.edit_case = data

                    st.session_state.selected_case = None

                    st.rerun()

            with _detail_gap2:

                st.markdown(
                    '<div class="detail-action-gap"></div>',
                    unsafe_allow_html=True,
                )

            with btn2:

                if st.button(
                    "🗑️ 삭제하기",
                    key="detail_delete_btn",
                    use_container_width=True
                ):

                    c.execute(
                        "DELETE FROM cases WHERE case_id = ?",
                        (data[0],)
                    )

                    conn.commit()

                    st.success("삭제 완료")

                    st.session_state.selected_case = None

                    st.rerun()

    # --------------------------
    # 수정화면
    # --------------------------

    def show_edit():

        data = st.session_state.edit_case

        if st.session_state.get("edit_date_case_id") != data[0]:
            try:
                st.session_state.edit_case_date = datetime.strptime(
                    data[0][:8],
                    "%Y%m%d"
                ).date()
            except ValueError:
                st.session_state.edit_case_date = datetime.now().date()

            st.session_state.edit_date_case_id = data[0]
            st.session_state.edit_date_edit_open = False

        edit_case_id_col, edit_date_col = st.columns([2.2, 1], gap="small")

        with edit_case_id_col:

            edit_preview_case_id = get_next_case_id(
                st.session_state.edit_case_date,
                current_case_id=data[0]
            )

            st.text(f"사례번호 : {edit_preview_case_id}")

        with edit_date_col:

            if st.button(
                "📅 일자 수정",
                key="edit_date_edit_btn"
            ):

                st.session_state.edit_date_edit_open = (
                    not st.session_state.edit_date_edit_open
                )

        if st.session_state.edit_date_edit_open:

            selected_edit_date = st.date_input(
                "발생일자",
                value=st.session_state.edit_case_date,
                key=f"edit_case_date_input_{data[0]}"
            )

            if selected_edit_date != st.session_state.edit_case_date:

                st.session_state.edit_case_date = selected_edit_date

                st.rerun()

        st.divider()

        category_list = [
            "파손",
            "오발송",
            "쇼트",
            "불량",
            "누락",
            "기타",
            "수평",
            "용접"
        ]

        edit_category_value = "누락" if data[1] == "변심" else data[1]

        category = st.selectbox(
            "유형",
            category_list,
            index=category_list.index(edit_category_value),
            key="edit_category"
        )

        product = st.text_input(
            "상품명",
            value=data[3],
            key="edit_product"
        )

        barcode = st.text_input(
            "바코드",
            value=data[2],
            key="edit_barcode"
        )

        st.divider()

        st.markdown("### 상품 대표사진")

        if data[8]:

            img = Image.open(BytesIO(data[8]))

            img = ImageOps.exif_transpose(img)

            st.image(
                img,
                width=250
            )

        product_image = st.file_uploader(
            "대표사진 변경",
            type=["jpg", "jpeg", "png"],
            key="edit_product_image"
        )

        st.divider()

        cause = st.text_area(
            "원인",
            value=data[4],
            height=120,
            key="edit_cause"
        )

        st.markdown("### 사례 사진")

        case_image = st.file_uploader(
            "사례사진 변경",
            type=["jpg", "jpeg", "png"],
            key="edit_case_image"
        )

        case_image_upload_data = case_image.getvalue() if case_image else None
        case_image_original_saved_data = get_case_image_original(data[0])
        case_image_saved_data = first_readable_image_data(
            data[9],
            case_image_original_saved_data,
        )
        case_image_current_data = first_readable_image_data(
            case_image_upload_data,
            case_image_saved_data,
        )
        case_image_original_data = first_readable_image_data(
            case_image_upload_data,
            case_image_original_saved_data,
            case_image_saved_data,
        )

        if case_image_current_data:

            img = image_bytes_to_rgb(case_image_current_data)

            st.image(
                img,
                width=150
            )

        st.session_state[f"edit_case_marker_{data[0]}_red_restore_image"] = (
            case_image_original_data
        )

        case_image_data = edit_red_circle_marker_tool(
            case_image_current_data,
            f"edit_case_marker_{data[0]}",
            "사례사진 부위 표시",
            restore_image_data=case_image_original_data,
        )

        st.divider()

        action = st.text_area(
            "조치방법",
            value=data[5],
            height=120,
            key="edit_action"
        )

        repair_method = st.text_area(
            "수리방법",
            value=data[6],
            height=120,
            key="edit_repair_method"
        )

        st.markdown("### 수리 사진")

        repair_image = st.file_uploader(
            "수리사진 변경",
            type=["jpg", "jpeg", "png"],
            key="edit_repair_image"
        )

        repair_image_upload_data = repair_image.getvalue() if repair_image else None
        repair_image_original_saved_data = get_repair_image_original(data[0])
        repair_image_saved_data = first_readable_image_data(
            data[10],
            repair_image_original_saved_data,
        )
        repair_image_current_data = first_readable_image_data(
            repair_image_upload_data,
            repair_image_saved_data,
        )
        repair_image_original_data = first_readable_image_data(
            repair_image_upload_data,
            repair_image_original_saved_data,
            repair_image_saved_data,
        )

        repair_marker_prefix = f"edit_repair_marker_{data[0]}"
        repair_current_fingerprint = (
            image_data_fingerprint(repair_image_current_data)
            if repair_image_current_data
            else None
        )
        repair_last_fingerprint_key = f"{repair_marker_prefix}_current_fingerprint"

        if (
            repair_current_fingerprint
            and st.session_state.get(repair_last_fingerprint_key) != repair_current_fingerprint
        ):
            clear_marker_session_state(repair_marker_prefix, include_canvas_state=True)
            restore_marker_canvas(repair_marker_prefix, "blue")
            st.session_state[repair_last_fingerprint_key] = repair_current_fingerprint

        st.session_state[f"{repair_marker_prefix}_blue_restore_image"] = (
            repair_image_original_data
        )

        repair_image_display_data = first_readable_image_data(
            repair_image_current_data,
            repair_image_original_data,
        )

        if repair_image_display_data:

            img = image_bytes_to_rgb(repair_image_display_data)

            st.image(
                img,
                width=150
            )

        repair_image_data = blue_circle_marker_tool(
            repair_image_display_data,
            repair_marker_prefix,
            restore_image_data=repair_image_original_data,
        )

        st.divider()

        st.markdown(
            '<div class="edit-prevention-top-gap"></div>',
            unsafe_allow_html=True
        )

        prevention = st.text_area(
            "방지대책",
            value=data[7],
            height=120,
            key="edit_prevention"
        )

        st.divider()

        st.markdown(
            '<div class="edit-button-top-gap"></div>',
            unsafe_allow_html=True
        )

        col1, col2 = st.columns(2)

        with col1:

            if st.button(
                "💾 수정 저장",
                key="edit_save_btn",
                use_container_width=True
            ):

                product_image_data = (
                    product_image.getvalue()
                    if product_image
                    else data[8]
                )

                new_case_id = get_next_case_id(
                    st.session_state.edit_case_date,
                    current_case_id=data[0]
                )

                c.execute("""
                UPDATE cases
                SET
                    case_id = ?,
                    category = ?,
                    barcode = ?,
                    product = ?,
                    cause = ?,
                    action = ?,
                    repair_method = ?,
                    prevention = ?,
                    product_image = ?,
                    case_image = ?,
                    case_image_original = ?,
                    repair_image = ?,
                    repair_image_original = ?
                WHERE case_id = ?
                """,
                (
                    new_case_id,
                    category,
                    barcode,
                    product,
                    cause,
                    action,
                    repair_method,
                    prevention,
                    product_image_data,
                    case_image_data,
                    case_image_original_data,
                    repair_image_data,
                    repair_image_original_data,
                    data[0]
                ))

                conn.commit()

                clear_marker_session_state(f"edit_case_marker_{data[0]}")
                clear_marker_session_state(f"edit_repair_marker_{data[0]}")

                c.execute("""
                SELECT
                    case_id,
                    category,
                    barcode,
                    product,
                    cause,
                    action,
                    repair_method,
                    prevention,
                    product_image,
                    case_image,
                    repair_image
                FROM cases
                WHERE case_id = ?
                """, (new_case_id,))

                st.session_state.selected_case = c.fetchone()

                st.session_state.edit_case = None
                st.session_state.edit_date_edit_open = False
                st.session_state.edit_date_case_id = None

                st.success("수정 완료")

                st.rerun()

        with col2:

            if st.button(
                "↩️ 취소",
                key="edit_cancel_btn",
                use_container_width=True
            ):

                clear_marker_session_state(f"edit_case_marker_{data[0]}")
                clear_marker_session_state(f"edit_repair_marker_{data[0]}")

                st.session_state.selected_case = data

                st.session_state.edit_case = None
                st.session_state.edit_date_edit_open = False
                st.session_state.edit_date_case_id = None

                st.rerun()

    # --------------------------
    # 오른쪽 화면
    # --------------------------

    with right:

        if st.session_state.edit_case:

            show_edit()

        elif st.session_state.selected_case:

            show_detail()

        else:

            show_dashboard()
    # ==========================
    # 엑셀 내보내기
    # ==========================

    def create_excel():

        c.execute("""
        SELECT
            case_id,
            category,
            barcode,
            product,
            cause,
            action,
            repair_method,
            prevention,
            product_image,
            case_image,
            repair_image
        FROM cases
        ORDER BY id DESC
        """)

        rows = c.fetchall()

        wb = Workbook()

        # =====================
        # 시트1 : 사례목록
        # =====================

        ws1 = wb.active
        ws1.title = "사례목록"

        ws1.append([
            "사례번호",
            "유형",
            "바코드",
            "상품명",
            "원인",
            "조치방법",
            "수리방법",
            "방지대책"
        ])

        for row in rows:

            ws1.append([
                row[0],
                row[1],
                row[2],
                row[3],
                row[4],
                row[5],
                row[6],
                row[7]
            ])

        # =====================
        # 시트2 : 사진목록
        # =====================

        ws2 = wb.create_sheet("사진목록")

        ws2.append([
            "사례번호",
            "대표사진",
            "사례사진",
            "수리사진"
        ])

        ws2.column_dimensions["A"].width = 18
        ws2.column_dimensions["B"].width = 12
        ws2.column_dimensions["C"].width = 12
        ws2.column_dimensions["D"].width = 12

        photo_row = 2
        temp_image_paths = []

        for row in rows:

            ws2.cell(
                row=photo_row,
                column=1,
                value=row[0]
            )

            image_data_list = [
                (row[8], "B"),
                (row[9], "C"),
                (row[10], "D")
            ]

            for img_data, col_letter in image_data_list:

                if img_data:

                    tmp = tempfile.NamedTemporaryFile(
                        delete=False,
                        suffix=".png"
                    )

                    tmp.write(img_data)
                    tmp.close()
                    temp_image_paths.append(Path(tmp.name))

                    img = XLImage(tmp.name)

                    # 썸네일 크기
                    img.width = 80
                    img.height = 80

                    ws2.add_image(
                        img,
                        f"{col_letter}{photo_row}"
                    )

            ws2.row_dimensions[photo_row].height = 65

            photo_row += 1

        excel_buffer = BytesIO()

        try:
            wb.save(excel_buffer)
        finally:
            for temp_image_path in temp_image_paths:
                try:
                    temp_image_path.unlink(missing_ok=True)
                except OSError:
                    pass

        excel_buffer.seek(0)

        return excel_buffer

    excel_export_key = "return_case_excel_export_bytes"

    if st.button("📥 엑셀 파일 생성", key="main_excel_prepare_btn"):
        with st.spinner("엑셀 파일을 생성하는 중입니다..."):
            st.session_state[excel_export_key] = create_excel().getvalue()
        st.success("엑셀 파일이 준비되었습니다.")

    if excel_export_key in st.session_state:
        st.download_button(
            label="📥 준비된 엑셀 다운로드",
            data=st.session_state[excel_export_key],
            file_name="반품사례DB.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="main_excel_download_btn"
        )


if __name__ == "__main__":
    st.set_page_config(
        page_title="LOGIN_반품/AS",
        layout="wide",
    )
    render_return_case_system()
